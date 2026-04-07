"""
按 Agent 角色构建外部工具上下文。

这个服务负责把“主 Agent 的命令”转换成真正可执行的工具上下文：
- 先做 command gate 决策
- 再按 Agent 类型选择具体工具
- 最后输出统一结构和审计日志
"""

from __future__ import annotations

import asyncio
import csv
from collections import deque
from datetime import datetime
from hashlib import sha1
import json
import os
from pathlib import Path
import re
import shutil
import sqlite3
import subprocess
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import urlsplit, urlunsplit

import structlog
try:
    import asyncpg
except Exception:  # pragma: no cover - optional dependency
    asyncpg = None  # type: ignore[assignment]

from app.config import settings
from app.models.tooling import AgentToolingConfig
from app.runtime.connectors import (
    APMConnector,
    AlertPlatformConnector,
    CMDBConnector,
    GrafanaConnector,
    LogCloudConnector,
    LokiConnector,
    PrometheusConnector,
    TelemetryConnector,
)
from app.services.code_analysis.call_graph_builder import (
    build_method_call_chain,
    parse_interface_ref,
    resolve_next_method_call,
)
from app.services.code_analysis.source_loader import (
    extract_field_types,
    extract_methods,
    find_source_unit,
    guess_entry_method,
    load_repo_focus_windows,
    load_source_units,
    parse_source_unit,
    resolve_repo_file,
)
from app.services.code_analysis.symbol_resolver import (
    expand_related_code_files,
    extract_related_code_symbols,
    find_symbol_file,
)
from app.services.tooling_service import tooling_service
from app.services.agent_skill_service import agent_skill_service
from app.services.mcp_service import mcp_service
from app.services.tool_plugin_gateway import ToolPluginGateway
from app.services.knowledge_service import knowledge_service
from app.services.tool_context.audit import ToolAuditBuilder
from app.services.tool_context.assemblers.change_focused import build_change_focused_context
from app.services.tool_context.assemblers.code_focused import build_code_focused_context
from app.services.tool_context.assemblers.cross_agent_focused import (
    build_cross_agent_focused_context,
    build_critique_summary,
    build_judge_verdict_summary,
    build_problem_coordination_summary,
    build_rebuttal_summary,
    build_rule_summary,
    build_verification_summary,
)
from app.services.tool_context.assemblers.database_focused import build_database_focused_context
from app.services.tool_context.assemblers.domain_focused import build_domain_focused_context
from app.services.tool_context.focused_context import resolve_focused_context_builder_name
from app.services.tool_context.assemblers.log_focused import build_log_focused_context
from app.services.tool_context.assemblers.metrics_focused import build_metrics_focused_context
from app.services.tool_context.assemblers.impact_focused import build_impact_focused_context
from app.services.tool_context.providers.change_provider import build_change_context as build_change_context_provider
from app.services.tool_context.providers.code_provider import build_code_context as build_code_context_provider
from app.services.tool_context.providers.database_provider import build_database_context as build_database_context_provider
from app.services.tool_context.providers.domain_provider import build_domain_context as build_domain_context_provider
from app.services.tool_context.providers.log_provider import build_log_context as build_log_context_provider
from app.services.tool_context.providers.metrics_provider import build_metrics_context as build_metrics_context_provider
from app.services.tool_context.providers.rule_suggestion_provider import (
    build_rule_suggestion_context as build_rule_suggestion_context_provider,
)
from app.services.tool_context.providers.runbook_provider import build_runbook_context as build_runbook_context_provider
from app.services.tool_context.assemblers.runbook_focused import build_runbook_focused_context
from app.services.tool_context.result import ToolContextResult
from app.services.tool_context.router import decide_tool_invocation, resolve_context_builder_name
from app.tools.case_library import CaseLibraryTool

logger = structlog.get_logger()


SOURCE_SUFFIXES = {
    ".py",
    ".java",
    ".kt",
    ".go",
    ".ts",
    ".tsx",
    ".js",
    ".jsx",
    ".rs",
    ".sql",
    ".yaml",
    ".yml",
    ".json",
    ".xml",
    ".properties",
    ".md",
}

GIT_FETCH_TIMEOUTS = (15, 25)
GIT_CLONE_TIMEOUTS = (30, 45)
GIT_LOCAL_TIMEOUT = 20

class AgentToolContextService:
    """封装AgentToolContextService相关数据结构或服务能力。"""
    def __init__(self) -> None:
        """初始化本地工具、远端连接器和审计计数器。"""
        self._case_library = CaseLibraryTool()
        self._telemetry_connector = TelemetryConnector()
        self._cmdb_connector = CMDBConnector()
        self._prometheus_connector = PrometheusConnector()
        self._loki_connector = LokiConnector()
        self._grafana_connector = GrafanaConnector()
        self._apm_connector = APMConnector()
        self._logcloud_connector = LogCloudConnector()
        self._alert_platform_connector = AlertPlatformConnector()
        # 中文注释：插件网关负责执行 extensions/tools 下的可扩展工具，不替代内置 provider。
        self._tool_plugin_gateway = ToolPluginGateway()
        self._audit_builder = ToolAuditBuilder()
        # 中文注释：仅专家类 Agent 默认尝试调用 MCP 服务，避免协调/裁决角色产生无效外部请求。
        self._mcp_expert_agents = {
            "LogAgent",
            "MetricsAgent",
            "DatabaseAgent",
            "DomainAgent",
            "CodeAgent",
            "ChangeAgent",
            "ImpactAnalysisAgent",
            "RunbookAgent",
            "RuleSuggestionAgent",
        }

    async def build_context(
        self,
        *,
        agent_name: str,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """
        构建某个 Agent 当前轮次可用的工具上下文。

        这是服务主入口，负责：
        1. 根据命令判断是否允许调用工具。
        2. 根据 Agent 角色路由到对应 `_build_*_context`。
        3. 把 skill 注入和工具审计并入同一份结果。
        """
        command_gate = self._decide_tool_invocation(agent_name=agent_name, assigned_command=assigned_command)
        cfg = await tooling_service.get_config()
        builder_name = resolve_context_builder_name(agent_name)
        if builder_name:
            builder = getattr(self, builder_name)
            if builder_name == "_build_runbook_context":
                result = await builder(compact_context, incident_context, assigned_command, command_gate)
            else:
                result = await builder(cfg, compact_context, incident_context, assigned_command, command_gate)
        else:
            result = ToolContextResult(
                name="none",
                enabled=False,
                used=False,
                status="skipped",
                summary="当前 Agent 无外部工具配置。",
                data={},
                command_gate=command_gate,
                audit_log=[
                    self._audit(
                        tool_name="none",
                        action="tool_skip",
                        status="skipped",
                        detail={"reason": "当前 Agent 无外部工具配置。"},
                    )
                ],
            )
        result = self._merge_skill_context(
            result=result,
            cfg=cfg,
            agent_name=agent_name,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
        )
        result = self._merge_plugin_tool_context(
            result=result,
            cfg=cfg,
            agent_name=agent_name,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
        )
        result = await self._merge_mcp_context(
            result=result,
            agent_name=agent_name,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
        )
        # 无论实际是否命中工具，最终都会把标准化 investigation leads 带回去，
        # 让 Agent 在工具不可用时仍可基于已有线索完成受限分析。
        result.data = {
            **dict(result.data or {}),
            "investigation_leads": self._extract_investigation_leads(
                compact_context,
                incident_context,
                assigned_command,
            ),
        }
        result.permission_decision = {
            "allow_tool": bool(command_gate.get("allow_tool")),
            "reason": str(command_gate.get("reason") or ""),
            "decision_source": str(command_gate.get("decision_source") or ""),
        }
        if not result.execution_path:
            if result.name in {"telemetry_connector", "cmdb_connector"}:
                result.execution_path = "remote"
            elif result.name in {
                "git_repo_search",
                "git_change_window",
                "local_log_reader",
                "domain_excel_lookup",
                "db_snapshot_reader",
            }:
                result.execution_path = "local"
            else:
                result.execution_path = "none"
        return result.to_dict()

    def build_focused_context(
        self,
        *,
        agent_name: str,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]] = None,
        assigned_command: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        """为指定 Agent 生成更贴近问题闭包的专属上下文。"""
        builder_name = resolve_focused_context_builder_name(agent_name)
        if builder_name:
            builder = getattr(self, builder_name)
            return builder(compact_context, incident_context, tool_context, assigned_command)
        return {}

    def _merge_skill_context(
        self,
        *,
        result: ToolContextResult,
        cfg: AgentToolingConfig,
        agent_name: str,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
    ) -> ToolContextResult:
        """把 skill 命中结果并入已有工具上下文，形成统一的注入视图。"""
        gate = dict(result.command_gate or {})
        if not bool(gate.get("has_command")):
            return result
        if bool(gate.get("has_command")) and not bool(gate.get("allow_tool")):
            return result

        skill_result = agent_skill_service.select_skills(
            agent_name=agent_name,
            cfg=cfg.skills,
            assigned_command=assigned_command,
            compact_context=compact_context,
            incident_context=incident_context,
        )
        if not bool(skill_result.get("enabled")):
            return result
        if not bool(skill_result.get("used")):
            return result

        skill_payload = {
            "status": str(skill_result.get("status") or ""),
            "summary": str(skill_result.get("summary") or ""),
            "items": list(skill_result.get("skills") or []),
        }
        combined_data = dict(result.data or {})
        combined_data["skill_context"] = skill_payload

        normalized_skill_audit: List[Dict[str, Any]] = []
        for entry in list(skill_result.get("audit_log") or []):
            if not isinstance(entry, dict):
                continue
            normalized_skill_audit.append(
                self._audit(
                    tool_name="agent_skill_router",
                    action=str(entry.get("action") or "skill_call"),
                    status=str(entry.get("status") or "ok"),
                    detail=dict(entry.get("detail") or {}),
                )
            )

        if result.name in {"none", ""}:
            return ToolContextResult(
                name="agent_skill_router",
                enabled=True,
                used=True,
                status="ok",
                summary=str(skill_result.get("summary") or "Skill 调用成功。"),
                data=combined_data,
                command_gate=dict(result.command_gate or {}),
                audit_log=[*list(result.audit_log or []), *normalized_skill_audit],
                execution_path="local",
                permission_decision=dict(result.permission_decision or {}),
            )

        if not bool(result.used):
            base_snapshot = {
                "name": result.name,
                "enabled": bool(result.enabled),
                "used": bool(result.used),
                "status": str(result.status or ""),
                "summary": str(result.summary or ""),
            }
            combined_data["base_tool_context"] = base_snapshot
            return ToolContextResult(
                name="agent_skill_router",
                enabled=True,
                used=True,
                status="ok",
                summary=(
                    f"{str(skill_result.get('summary') or '').strip()}；"
                    f"原工具状态={base_snapshot['status'] or 'unknown'}"
                ).strip("；"),
                data=combined_data,
                command_gate=dict(result.command_gate or {}),
                audit_log=[*list(result.audit_log or []), *normalized_skill_audit],
                execution_path="local",
                permission_decision=dict(result.permission_decision or {}),
            )

        result.data = combined_data
        result.summary = f"{result.summary}；{str(skill_result.get('summary') or '').strip()}".strip("；")
        result.audit_log = [*list(result.audit_log or []), *normalized_skill_audit]
        if not result.execution_path:
            result.execution_path = "local"
        return result

    async def _merge_mcp_context(
        self,
        *,
        result: ToolContextResult,
        agent_name: str,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
    ) -> ToolContextResult:
        """将 MCP 取证结果并入工具上下文。"""
        if agent_name not in self._mcp_expert_agents:
            return result
        gate = dict(result.command_gate or {})
        if not bool(gate.get("has_command")) or not bool(gate.get("allow_tool")):
            return result

        mcp_result = await mcp_service.collect_agent_evidence(
            agent_name=agent_name,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
        )
        if not bool(mcp_result.get("enabled")):
            return result

        combined_data = dict(result.data or {})
        combined_data["mcp_context"] = {
            "summary": str(mcp_result.get("summary") or ""),
            "servers": list(mcp_result.get("servers") or []),
            "items": list(mcp_result.get("items") or []),
        }
        mcp_audit = [
            self._audit(
                tool_name="mcp_gateway",
                action=str(item.get("action") or "mcp_fetch"),
                status=str(item.get("status") or "ok"),
                detail=dict(item.get("detail") or {}),
            )
            for item in list(mcp_result.get("audit_log") or [])
            if isinstance(item, dict)
        ]

        mcp_summary = str(mcp_result.get("summary") or "MCP 取证完成。")
        if result.name in {"none", ""}:
            return ToolContextResult(
                name="mcp_gateway",
                enabled=True,
                used=bool(mcp_result.get("used")),
                status="ok" if bool(mcp_result.get("used")) else "skipped",
                summary=mcp_summary,
                data=combined_data,
                command_gate=dict(result.command_gate or {}),
                audit_log=[*list(result.audit_log or []), *mcp_audit],
                execution_path="remote",
                permission_decision=dict(result.permission_decision or {}),
            )

        result.data = combined_data
        result.summary = f"{str(result.summary or '').strip()}；{mcp_summary}".strip("；")
        result.audit_log = [*list(result.audit_log or []), *mcp_audit]
        if bool(mcp_result.get("used")):
            result.used = True
            if result.execution_path == "none":
                result.execution_path = "remote"
        return result

    def _merge_plugin_tool_context(
        self,
        *,
        result: ToolContextResult,
        cfg: AgentToolingConfig,
        agent_name: str,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
    ) -> ToolContextResult:
        """把插件 Tool 执行结果并入当前工具上下文。"""
        plugin_cfg = getattr(cfg, "tool_plugins", None)
        if not plugin_cfg or not bool(getattr(plugin_cfg, "enabled", False)):
            return result
        gate = dict(result.command_gate or {})
        if not bool(gate.get("has_command")) or not bool(gate.get("allow_tool")):
            return result

        requested_tools = self._collect_plugin_tool_requests(result=result, assigned_command=assigned_command)
        if not requested_tools:
            return result

        payload = {
            "agent_name": agent_name,
            "compact_context": compact_context,
            "incident_context": incident_context,
            "assigned_command": dict(assigned_command or {}),
            "tool_context": {
                "name": str(result.name or ""),
                "status": str(result.status or ""),
                "summary": str(result.summary or ""),
                "data": dict(result.data or {}),
            },
        }
        outputs = self._tool_plugin_gateway.invoke_for_agent(
            agent_name=agent_name,
            requested_tools=requested_tools,
            payload=payload,
            plugins_dir=str(getattr(plugin_cfg, "plugins_dir", "") or ""),
            timeout_seconds=int(getattr(plugin_cfg, "default_timeout_seconds", 60) or 60),
            allowlist=list(getattr(plugin_cfg, "allowed_tools", []) or []),
            max_calls=int(getattr(plugin_cfg, "max_calls", 3) or 3),
        )
        if not outputs:
            return result

        combined_data = dict(result.data or {})
        combined_data["plugin_tool_outputs"] = outputs
        plugin_audit: List[Dict[str, Any]] = []
        for item in outputs:
            success = bool(item.get("success"))
            status = "ok" if success else ("timeout" if bool(item.get("timed_out")) else "failed")
            plugin_audit.append(
                self._audit(
                    tool_name="agent_tool_plugin_router",
                    action="plugin_tool_call",
                    status=status,
                    detail={
                        "tool_name": str(item.get("tool_name") or ""),
                        "summary": str(item.get("summary") or "")[:220],
                        "requested_tools": requested_tools,
                    },
                )
            )

        plugin_summary = f"扩展工具命中 {len(outputs)} 个：{', '.join(str(item.get('tool_name') or '') for item in outputs)}"
        if result.name in {"none", ""}:
            return ToolContextResult(
                name="agent_tool_plugin_router",
                enabled=True,
                used=True,
                status="ok",
                summary=plugin_summary,
                data=combined_data,
                command_gate=dict(result.command_gate or {}),
                audit_log=[*list(result.audit_log or []), *plugin_audit],
                execution_path="local",
                permission_decision=dict(result.permission_decision or {}),
            )

        if not bool(result.used):
            base_snapshot = {
                "name": result.name,
                "enabled": bool(result.enabled),
                "used": bool(result.used),
                "status": str(result.status or ""),
                "summary": str(result.summary or ""),
            }
            combined_data["base_tool_context"] = base_snapshot
            return ToolContextResult(
                name="agent_tool_plugin_router",
                enabled=True,
                used=True,
                status="ok",
                summary=f"{plugin_summary}；原工具状态={base_snapshot['status'] or 'unknown'}",
                data=combined_data,
                command_gate=dict(result.command_gate or {}),
                audit_log=[*list(result.audit_log or []), *plugin_audit],
                execution_path="local",
                permission_decision=dict(result.permission_decision or {}),
            )

        result.data = combined_data
        result.summary = f"{str(result.summary or '').strip()}；{plugin_summary}".strip("；")
        result.audit_log = [*list(result.audit_log or []), *plugin_audit]
        if not result.execution_path:
            result.execution_path = "local"
        return result

    @staticmethod
    def _collect_plugin_tool_requests(
        *,
        result: ToolContextResult,
        assigned_command: Optional[Dict[str, Any]],
    ) -> List[str]:
        """收集本轮需要调用的插件 Tool 列表。"""
        command = dict(assigned_command or {})
        picks: List[str] = []
        raw_tool_hints = command.get("tool_hints")
        if isinstance(raw_tool_hints, list):
            for item in raw_tool_hints:
                name = str(item or "").strip()
                if name:
                    picks.append(name)

        skill_items = list((((result.data or {}).get("skill_context") or {}).get("items") or []))
        for skill in skill_items:
            if not isinstance(skill, dict):
                continue
            required_tools = skill.get("required_tools")
            if not isinstance(required_tools, list):
                continue
            for tool_name in required_tools:
                name = str(tool_name or "").strip()
                if name:
                    picks.append(name)

        return list(dict.fromkeys(picks))

    def _build_cross_agent_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        return build_cross_agent_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    def _build_problem_coordination_summary(
        self,
        *,
        problem_frame: Dict[str, Any],
        investigation_focus: Dict[str, Any],
        tool_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        return build_problem_coordination_summary(
            problem_frame=problem_frame,
            investigation_focus=investigation_focus,
            tool_summary=tool_summary,
        )

    def _build_judge_verdict_summary(
        self,
        *,
        problem_frame: Dict[str, Any],
        investigation_focus: Dict[str, Any],
        tool_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        return build_judge_verdict_summary(
            problem_frame=problem_frame,
            investigation_focus=investigation_focus,
            tool_summary=tool_summary,
        )

    def _build_verification_summary(
        self,
        *,
        problem_frame: Dict[str, Any],
        investigation_focus: Dict[str, Any],
        tool_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        return build_verification_summary(
            problem_frame=problem_frame,
            investigation_focus=investigation_focus,
            tool_summary=tool_summary,
        )

    def _build_critique_summary(
        self,
        *,
        problem_frame: Dict[str, Any],
        investigation_focus: Dict[str, Any],
        tool_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        return build_critique_summary(
            problem_frame=problem_frame,
            investigation_focus=investigation_focus,
            tool_summary=tool_summary,
        )

    def _build_rebuttal_summary(
        self,
        *,
        problem_frame: Dict[str, Any],
        investigation_focus: Dict[str, Any],
        tool_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        return build_rebuttal_summary(
            problem_frame=problem_frame,
            investigation_focus=investigation_focus,
            tool_summary=tool_summary,
        )

    def _build_rule_summary(
        self,
        *,
        problem_frame: Dict[str, Any],
        investigation_focus: Dict[str, Any],
        tool_summary: Dict[str, Any],
    ) -> Dict[str, Any]:
        return build_rule_summary(
            problem_frame=problem_frame,
            investigation_focus=investigation_focus,
            tool_summary=tool_summary,
        )

    def _build_code_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        return build_code_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    def _build_log_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        return build_log_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    def _build_domain_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        return build_domain_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    def _build_database_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        return build_database_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    def _build_metrics_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        return build_metrics_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    def _build_impact_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        # 中文注释：影响面专家需要同时看到“功能层”和“接口层”的线索，
        # 因此这里把 incident、责任田映射、已有调查线索统一压缩成同一份 focused context。
        return build_impact_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    def _build_change_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        return build_change_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    def _build_runbook_focused_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        return build_runbook_focused_context(
            self,
            compact_context,
            incident_context,
            tool_context,
            assigned_command,
        )

    async def _build_code_context(
        self,
        cfg: AgentToolingConfig,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        command_gate: Dict[str, Any],
    ) -> ToolContextResult:
        """构建构建代码上下文，供后续节点或调用方直接使用。"""
        return await build_code_context_provider(
            self,
            cfg=cfg,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
            command_gate=command_gate,
        )

    async def _build_log_context(
        self,
        cfg: AgentToolingConfig,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        command_gate: Dict[str, Any],
    ) -> ToolContextResult:
        """构建构建日志上下文，供后续节点或调用方直接使用。"""
        return await build_log_context_provider(
            self,
            cfg=cfg,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
            command_gate=command_gate,
        )

    async def _build_domain_context(
        self,
        cfg: AgentToolingConfig,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        command_gate: Dict[str, Any],
    ) -> ToolContextResult:
        """构建构建domain上下文，供后续节点或调用方直接使用。"""
        return await build_domain_context_provider(
            self,
            cfg=cfg,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
            command_gate=command_gate,
        )

    async def _build_metrics_context(
        self,
        cfg: AgentToolingConfig,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        command_gate: Dict[str, Any],
    ) -> ToolContextResult:
        """构建构建metrics上下文，供后续节点或调用方直接使用。"""
        return await build_metrics_context_provider(
            self,
            cfg=cfg,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
            command_gate=command_gate,
        )

    async def _build_change_context(
        self,
        cfg: AgentToolingConfig,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        command_gate: Dict[str, Any],
    ) -> ToolContextResult:
        """构建构建change上下文，供后续节点或调用方直接使用。"""
        return await build_change_context_provider(
            self,
            cfg=cfg,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
            command_gate=command_gate,
        )

    async def _build_database_context(
        self,
        cfg: AgentToolingConfig,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        command_gate: Dict[str, Any],
    ) -> ToolContextResult:
        """构建构建database上下文，供后续节点或调用方直接使用。"""
        return await build_database_context_provider(
            self,
            cfg=cfg,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
            command_gate=command_gate,
        )

    async def _build_runbook_context(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        command_gate: Dict[str, Any],
    ) -> ToolContextResult:
        """构建构建runbook上下文，供后续节点或调用方直接使用。"""
        return await build_runbook_context_provider(
            self,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
            command_gate=command_gate,
        )

    async def _build_rule_suggestion_context(
        self,
        cfg: AgentToolingConfig,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        command_gate: Dict[str, Any],
    ) -> ToolContextResult:
        """构建构建rulesuggestion上下文，供后续节点或调用方直接使用。"""
        return await build_rule_suggestion_context_provider(
            self,
            cfg=cfg,
            compact_context=compact_context,
            incident_context=incident_context,
            assigned_command=assigned_command,
            command_gate=command_gate,
        )

    def _resolve_repo_path(
        self,
        repo_url: str,
        access_token: str,
        branch: str,
        local_repo_path: str,
        audit_log: List[Dict[str, Any]],
    ) -> str:
        """执行resolverepopath相关逻辑，并为当前模块提供可复用的处理能力。"""
        raw_local_path = str(local_repo_path or "").strip()
        local_path = Path(raw_local_path) if raw_local_path else None
        if local_path and local_path.exists() and local_path.is_dir():
            audit_log.append(
                self._audit(
                    tool_name="git_repo_search",
                    action="repo_path_resolve",
                    status="ok",
                    detail={
                        "mode": "local",
                        "local_repo_path": str(local_path),
                    },
                )
            )
            return str(local_path)

        url = str(repo_url or "").strip()
        if not url:
            audit_log.append(
                self._audit(
                    tool_name="git_repo_search",
                    action="repo_path_resolve",
                    status="unavailable",
                    detail={"reason": "repo_url 为空且 local_repo_path 不可用"},
                )
            )
            return ""
        if not self._is_allowed_git_host(url):
            safe_url = self._mask_url_secret(url)
            audit_log.append(
                self._audit(
                    tool_name="git_repo_search",
                    action="permission_check",
                    status="denied",
                    detail={
                        "reason": "repo host not in allowlist",
                        "repo_url": safe_url,
                        "allowlist": list(settings.TOOL_GIT_HOST_ALLOWLIST or []),
                    },
                )
            )
            return ""

        cache_root = Path(settings.LOCAL_STORE_DIR) / "tool_cache" / "repos"
        cache_root.mkdir(parents=True, exist_ok=True)
        repo_key = sha1(url.encode("utf-8")).hexdigest()[:20]
        repo_path = cache_root / repo_key
        auth_url = self._inject_token(url, access_token)
        safe_branch = str(branch or "main").strip() or "main"
        safe_url = self._mask_url_secret(url)
        audit_log.append(
            self._audit(
                tool_name="git_repo_search",
                action="repo_path_resolve",
                status="ok",
                detail={
                    "mode": "remote",
                    "repo_url": safe_url,
                    "branch": safe_branch,
                    "cache_repo_path": str(repo_path),
                },
            )
        )

        if (repo_path / ".git").exists():
            audit_log.append(
                self._audit(
                    tool_name="git_repo_search",
                    action="http_request",
                    status="started",
                    detail={
                        "operation": "git_fetch",
                        "repo_url": safe_url,
                        "branch": safe_branch,
                    },
                )
            )
            try:
                self._run_git_with_retry(
                    ["git", "fetch", "--depth", "1", "origin", safe_branch],
                    cwd=repo_path,
                    audit_log=audit_log,
                    action="git_fetch",
                    repo_url=safe_url,
                    timeout_plan=GIT_FETCH_TIMEOUTS,
                )
                self._run_git(
                    ["git", "checkout", "-B", safe_branch, "FETCH_HEAD"],
                    cwd=repo_path,
                    audit_log=audit_log,
                    action="git_checkout",
                    repo_url="",
                    timeout_seconds=GIT_LOCAL_TIMEOUT,
                )
                self._run_git(
                    ["git", "reset", "--hard", "FETCH_HEAD"],
                    cwd=repo_path,
                    audit_log=audit_log,
                    action="git_reset",
                    repo_url="",
                    timeout_seconds=GIT_LOCAL_TIMEOUT,
                )
            except Exception as exc:
                error_text = str(exc).strip() or exc.__class__.__name__
                audit_log.append(
                    self._audit(
                        tool_name="git_repo_search",
                        action="repo_sync_degraded",
                        status="fallback",
                        detail={
                            "reason": "remote fetch 失败，回退到本地缓存仓库",
                            "repo_path": str(repo_path),
                            "error": error_text[:400],
                        },
                    )
                )
                logger.warning(
                    "git_repo_sync_degraded",
                    repo_path=str(repo_path),
                    repo_url=safe_url,
                    error=error_text[:400],
                )
            return str(repo_path)

        audit_log.append(
            self._audit(
                tool_name="git_repo_search",
                action="http_request",
                status="started",
                detail={
                    "operation": "git_clone",
                    "repo_url": safe_url,
                    "branch": safe_branch,
                },
            )
        )
        if repo_path.exists() and not (repo_path / ".git").exists():
            shutil.rmtree(repo_path, ignore_errors=True)
            audit_log.append(
                self._audit(
                    tool_name="git_repo_search",
                    action="repo_path_cleanup",
                    status="ok",
                    detail={
                        "reason": "清理不完整的历史 clone 目录",
                        "repo_path": str(repo_path),
                    },
                )
            )
        self._run_git_with_retry(
            [
                "git",
                "clone",
                "--depth",
                "1",
                "--filter=blob:none",
                "--single-branch",
                "--branch",
                safe_branch,
                auth_url,
                str(repo_path),
            ],
            cwd=cache_root,
            audit_log=audit_log,
            action="git_clone",
            repo_url=safe_url,
            timeout_plan=GIT_CLONE_TIMEOUTS,
        )
        return str(repo_path)

    def _run_git(
        self,
        cmd: List[str],
        cwd: Path,
        *,
        audit_log: List[Dict[str, Any]],
        action: str,
        repo_url: str,
        timeout_seconds: int,
    ) -> None:
        """负责运行git，并处理调用过程中的超时、错误与返回结果。"""
        started = datetime.utcnow()
        safe_cmd = [self._sanitize_command_part(item) for item in cmd]
        try:
            proc = subprocess.run(
                cmd,
                cwd=str(cwd),
                capture_output=True,
                text=True,
                timeout=max(10, int(timeout_seconds)),
                check=False,
                env=self._git_env(),
            )
        except subprocess.TimeoutExpired as exc:
            duration_ms = int((datetime.utcnow() - started).total_seconds() * 1000)
            audit_log.append(
                self._audit(
                    tool_name="git_repo_search",
                    action="git_command",
                    status="timeout",
                    detail={
                        "action": action,
                        "cwd": str(cwd),
                        "command": " ".join(safe_cmd),
                        "repo_url": repo_url,
                        "timeout_seconds": int(timeout_seconds),
                        "duration_ms": duration_ms,
                    },
                )
            )
            logger.warning(
                "tool_git_command_timeout",
                action=action,
                cwd=str(cwd),
                command=" ".join(safe_cmd),
                repo_url=repo_url,
                timeout_seconds=int(timeout_seconds),
                duration_ms=duration_ms,
            )
            raise RuntimeError(
                f"git 命令超时({int(timeout_seconds)}s): {' '.join(safe_cmd)}"
            ) from exc

        duration_ms = int((datetime.utcnow() - started).total_seconds() * 1000)
        detail = {
            "action": action,
            "cwd": str(cwd),
            "command": " ".join(safe_cmd),
            "repo_url": repo_url,
            "return_code": proc.returncode,
            "duration_ms": duration_ms,
            "stdout_preview": str(proc.stdout or "").strip()[:300],
            "stderr_preview": str(proc.stderr or "").strip()[:300],
        }
        audit_log.append(
            self._audit(
                tool_name="git_repo_search",
                action="git_command",
                status="ok" if proc.returncode == 0 else "error",
                detail=detail,
            )
        )
        logger.info(
            "tool_git_command",
            action=action,
            cwd=str(cwd),
            command=" ".join(safe_cmd),
            repo_url=repo_url,
            return_code=proc.returncode,
            duration_ms=duration_ms,
            stdout_preview=str(proc.stdout or "").strip()[:120],
            stderr_preview=str(proc.stderr or "").strip()[:120],
        )
        if proc.returncode != 0:
            raise RuntimeError(proc.stderr.strip() or proc.stdout.strip() or "git command failed")

    def _run_git_with_retry(
        self,
        cmd: List[str],
        *,
        cwd: Path,
        audit_log: List[Dict[str, Any]],
        action: str,
        repo_url: str,
        timeout_plan: tuple[int, ...],
    ) -> None:
        """负责运行gitwithretry，并处理调用过程中的超时、错误与返回结果。"""
        last_error: Optional[Exception] = None
        for index, timeout_seconds in enumerate(timeout_plan, start=1):
            try:
                self._run_git(
                    cmd,
                    cwd=cwd,
                    audit_log=audit_log,
                    action=action,
                    repo_url=repo_url,
                    timeout_seconds=int(timeout_seconds),
                )
                return
            except Exception as exc:
                last_error = exc
                audit_log.append(
                    self._audit(
                        tool_name="git_repo_search",
                        action="git_retry",
                        status="retrying" if index < len(timeout_plan) else "failed",
                        detail={
                            "operation": action,
                            "attempt": index,
                            "max_attempts": len(timeout_plan),
                            "timeout_seconds": int(timeout_seconds),
                            "error": str(exc)[:400],
                        },
                    )
                )
                if index < len(timeout_plan):
                    continue
        raise RuntimeError(str(last_error) if last_error else f"{action} failed")

    def _git_env(self) -> Dict[str, str]:
        """执行gitenv相关逻辑，并为当前模块提供可复用的处理能力。"""
        env = dict(os.environ)
        env.setdefault("GIT_TERMINAL_PROMPT", "0")
        env.setdefault("GIT_ASKPASS", "echo")
        return env

    def _is_allowed_git_host(self, repo_url: str) -> bool:
        """执行isallowedgithost相关逻辑，并为当前模块提供可复用的处理能力。"""
        raw = str(repo_url or "").strip()
        if not raw:
            return False
        parts = urlsplit(raw)
        host = str(parts.hostname or "").strip().lower()
        if not host:
            return False
        allowlist = [str(item or "").strip().lower() for item in (settings.TOOL_GIT_HOST_ALLOWLIST or []) if str(item or "").strip()]
        if not allowlist:
            return True
        return host in allowlist

    def _inject_token(self, repo_url: str, token: str) -> str:
        """执行injecttoken相关逻辑，并为当前模块提供可复用的处理能力。"""
        raw = str(repo_url or "").strip()
        tk = str(token or "").strip()
        if not tk:
            return raw
        parts = urlsplit(raw)
        if parts.scheme not in {"http", "https"}:
            return raw
        if "@" in parts.netloc:
            return raw
        netloc = f"oauth2:{tk}@{parts.netloc}"
        return urlunsplit((parts.scheme, netloc, parts.path, parts.query, parts.fragment))

    def _collect_recent_git_changes(
        self,
        repo_path: str,
        max_items: int,
        audit_log: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """执行收集recentgitchanges相关逻辑，并为当前模块提供可复用的处理能力。"""
        cmd = [
            "git",
            "--no-pager",
            "log",
            f"-n{max(1, min(int(max_items or 20), 80))}",
            "--pretty=format:%H\t%ad\t%an\t%s",
            "--date=iso",
        ]
        try:
            completed = subprocess.run(
                cmd,
                cwd=repo_path,
                capture_output=True,
                text=True,
                timeout=GIT_LOCAL_TIMEOUT,
                check=True,
                env=self._git_env(),
            )
            output = completed.stdout or ""
        except subprocess.CalledProcessError as exc:
            err = str(exc.stderr or exc.stdout or "").strip()
            err_lower = err.lower()
            no_commit_markers = (
                "does not have any commits yet",
                "your current branch",
                "no commits yet",
                "bad revision",
            )
            if any(marker in err_lower for marker in no_commit_markers):
                audit_log.append(
                    self._audit(
                        tool_name="git_change_window",
                        action="git_log_changes",
                        status="unavailable",
                        detail={
                            "repo_path": str(repo_path),
                            "reason": "仓库暂无可用提交记录",
                            "stderr": err[:300],
                        },
                    )
                )
                return []
            raise RuntimeError(err or "git log failed") from exc
        except subprocess.TimeoutExpired as exc:
            raise RuntimeError(f"git log timeout({GIT_LOCAL_TIMEOUT}s): {str(exc)[:200]}") from exc
        audit_log.append(
            self._audit(
                tool_name="git_change_window",
                action="git_log_changes",
                status="ok",
                detail={"repo_path": str(repo_path), "lines": len(output.splitlines())},
            )
        )
        changes: List[Dict[str, Any]] = []
        for line in output.splitlines():
            parts = line.split("\t")
            if len(parts) < 4:
                continue
            commit, commit_time, author, subject = parts[0], parts[1], parts[2], parts[3]
            changes.append(
                {
                    "commit": commit[:12],
                    "time": commit_time,
                    "author": author,
                    "subject": subject[:240],
                }
            )
        return changes

    def _collect_metrics_signals(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
    ) -> List[Dict[str, Any]]:
        """执行收集metrics信号相关逻辑，并为当前模块提供可复用的处理能力。"""
        signals: List[Dict[str, Any]] = []
        text_sources = [
            str(compact_context.get("log_excerpt") or ""),
            str(incident_context.get("log_content") or ""),
            str(incident_context.get("description") or ""),
            str(incident_context.get("remote_telemetry_payload") or ""),
            str(incident_context.get("remote_prometheus_payload") or ""),
            str(incident_context.get("remote_loki_payload") or ""),
        ]
        metric_patterns = [
            ("cpu", r"cpu[^0-9]*([0-9]+(?:\.[0-9]+)?%?)", "CPU"),
            ("threads", r"(?:线程|threads?)[^0-9]*([0-9]+)", "线程"),
            ("hikari_pending", r"hikari[^,\n]*pending[^0-9]*([0-9]+)", "Hikari Pending"),
            (
                "db_conn",
                r"(?:db_conn|db\.active\.connections|database\s+connections?)\s*[=:]?\s*([0-9]+/[0-9]+)",
                "DB连接",
            ),
            ("error_rate", r"(?:5xx|error(?:_rate)?)[^0-9]*([0-9]+(?:\.[0-9]+)?%?)", "错误率"),
        ]
        for source_text in text_sources:
            if not source_text:
                continue
            for metric_key, pattern, label in metric_patterns:
                for match in re.finditer(pattern, source_text, flags=re.IGNORECASE):
                    value = str(match.group(1) or "").strip()
                    if not value:
                        continue
                    start = max(0, match.start() - 50)
                    end = min(len(source_text), match.end() + 50)
                    snippet = source_text[start:end].strip()
                    signals.append(
                        {
                            "metric": metric_key,
                            "label": label,
                            "value": value,
                            "snippet": snippet[:280],
                        }
                    )
        dedup: List[Dict[str, Any]] = []
        seen = set()
        for item in signals:
            key = f"{item.get('metric')}|{item.get('value')}|{item.get('snippet')}"
            if key in seen:
                continue
            seen.add(key)
            dedup.append(item)
        return dedup[:40]

    async def _collect_postgres_snapshot(
        self,
        *,
        dsn: str,
        schema: str,
        max_rows: int,
        keywords: List[str],
        target_tables: List[str],
        timeout_seconds: int,
    ) -> Dict[str, Any]:
        """执行收集postgressnapshot相关逻辑，并为当前模块提供可复用的处理能力。"""
        conn = await asyncpg.connect(dsn=dsn, timeout=timeout_seconds)  # type: ignore[union-attr]
        try:
            table_records = await conn.fetch(
                """
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = $1 AND table_type = 'BASE TABLE'
                ORDER BY table_name
                """,
                schema,
            )
            all_table_rows = [str(row["table_name"]) for row in table_records]
            normalized_targets = {self._normalize_table_name(value) for value in (target_tables or []) if self._normalize_table_name(value)}
            table_rows = [name for name in all_table_rows if self._normalize_table_name(name) in normalized_targets] if normalized_targets else list(all_table_rows)
            used_target_tables = bool(normalized_targets)
            fallback_reason = ""
            if used_target_tables and not table_rows:
                table_rows = list(all_table_rows)
                fallback_reason = "mapped_tables_not_found_fallback_all"

            table_structures: List[Dict[str, Any]] = []
            indexes: Dict[str, List[Dict[str, Any]]] = {}
            for table_name in table_rows[:50]:
                column_rows = await conn.fetch(
                    """
                    SELECT column_name, data_type, is_nullable, column_default, ordinal_position
                    FROM information_schema.columns
                    WHERE table_schema = $1 AND table_name = $2
                    ORDER BY ordinal_position
                    """,
                    schema,
                    table_name,
                )
                columns = [
                    {
                        "name": str(row["column_name"]),
                        "type": str(row["data_type"] or ""),
                        "notnull": str(row["is_nullable"] or "").upper() == "NO",
                        "default": row["column_default"],
                    }
                    for row in column_rows
                ]
                table_structures.append({"table": table_name, "columns": columns})

                index_rows = await conn.fetch(
                    """
                    SELECT indexname, indexdef
                    FROM pg_indexes
                    WHERE schemaname = $1 AND tablename = $2
                    ORDER BY indexname
                    """,
                    schema,
                    table_name,
                )
                indexes[table_name] = [
                    {
                        "index": str(row["indexname"]),
                        "definition": str(row["indexdef"] or ""),
                        "unique": " UNIQUE " in str(row["indexdef"] or "").upper(),
                    }
                    for row in index_rows
                ]

            slow_sql = await self._pg_fetch_rows(
                conn,
                [
                    "SELECT query, calls, total_exec_time, mean_exec_time, rows FROM pg_stat_statements ORDER BY total_exec_time DESC LIMIT $1",
                    "SELECT query, calls, mean_exec_time FROM pg_stat_statements ORDER BY mean_exec_time DESC LIMIT $1",
                ],
                max_rows,
            )
            top_sql = await self._pg_fetch_rows(
                conn,
                [
                    "SELECT query, calls, total_exec_time, mean_exec_time, rows FROM pg_stat_statements ORDER BY calls DESC LIMIT $1",
                    "SELECT query, calls FROM pg_stat_statements ORDER BY calls DESC LIMIT $1",
                ],
                max_rows,
            )
            session_status = await self._pg_fetch_rows(
                conn,
                [
                    """
                    SELECT COALESCE(state, 'unknown') AS state,
                           COALESCE(wait_event_type, '') AS wait_event_type,
                           COALESCE(wait_event, '') AS wait_event,
                           COUNT(*)::int AS sessions
                    FROM pg_stat_activity
                    GROUP BY state, wait_event_type, wait_event
                    ORDER BY sessions DESC
                    LIMIT $1
                    """,
                    """
                    SELECT COALESCE(state, 'unknown') AS state,
                           COUNT(*)::int AS sessions
                    FROM pg_stat_activity
                    GROUP BY state
                    ORDER BY sessions DESC
                    LIMIT $1
                    """,
                ],
                max_rows,
            )

            keyword_hits: List[Dict[str, Any]] = []
            lowered_keywords = [str(word or "").lower().strip() for word in (keywords or []) if str(word or "").strip()]
            if lowered_keywords:
                for row in (slow_sql + top_sql)[:200]:
                    text = json.dumps(row, ensure_ascii=False).lower()
                    if any(k and k in text for k in lowered_keywords):
                        keyword_hits.append(row)
                        if len(keyword_hits) >= max_rows:
                            break

            return {
                "engine": "postgresql",
                "schema": schema,
                "requested_tables": list(target_tables or [])[:20],
                "used_target_tables": used_target_tables,
                "fallback_reason": fallback_reason,
                "total_table_count": len(all_table_rows),
                "table_count": len(table_rows),
                "tables": table_rows[:80],
                "table_structures": table_structures[:20],
                "indexes": indexes,
                "slow_sql": slow_sql[:max_rows],
                "top_sql": top_sql[:max_rows],
                "session_status": session_status[:max_rows],
                "keyword_hits": keyword_hits[:max_rows],
            }
        finally:
            await conn.close()

    def _collect_database_snapshot(
        self,
        db_path: Path,
        max_rows: int,
        keywords: List[str],
        target_tables: List[str],
    ) -> Dict[str, Any]:
        """执行收集databasesnapshot相关逻辑，并为当前模块提供可复用的处理能力。"""
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        try:
            cur = conn.cursor()
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
            all_table_rows = [str(row["name"]) for row in cur.fetchall()]
            normalized_targets = {self._normalize_table_name(value) for value in (target_tables or []) if self._normalize_table_name(value)}
            table_rows = [name for name in all_table_rows if self._normalize_table_name(name) in normalized_targets] if normalized_targets else list(all_table_rows)
            used_target_tables = bool(normalized_targets)
            fallback_reason = ""
            if used_target_tables and not table_rows:
                table_rows = list(all_table_rows)
                fallback_reason = "mapped_tables_not_found_fallback_all"

            table_structures: List[Dict[str, Any]] = []
            indexes: Dict[str, List[Dict[str, Any]]] = {}
            for table_name in table_rows[:50]:
                safe = self._escape_sql_identifier(table_name)
                cur.execute(f"PRAGMA table_info('{safe}')")
                columns = [
                    {
                        "name": str(r["name"]),
                        "type": str(r["type"] or ""),
                        "notnull": bool(r["notnull"]),
                        "default": r["dflt_value"],
                        "pk": bool(r["pk"]),
                    }
                    for r in cur.fetchall()
                ]
                table_structures.append({"table": table_name, "columns": columns})
                cur.execute(f"PRAGMA index_list('{safe}')")
                idx_rows = []
                for idx in cur.fetchall():
                    idx_name = str(idx["name"])
                    unique = bool(idx["unique"])
                    cur.execute(f"PRAGMA index_info('{self._escape_sql_identifier(idx_name)}')")
                    cols = [str(x["name"]) for x in cur.fetchall()]
                    idx_rows.append({"index": idx_name, "unique": unique, "columns": cols})
                indexes[table_name] = idx_rows

            slow_sql = self._query_first_existing(
                cur,
                [
                    "SELECT * FROM slow_sql ORDER BY duration_ms DESC LIMIT ?",
                    "SELECT * FROM slow_sql ORDER BY elapsed_ms DESC LIMIT ?",
                    "SELECT * FROM slow_sql ORDER BY cost_ms DESC LIMIT ?",
                    "SELECT * FROM slow_sql LIMIT ?",
                    "SELECT * FROM t_slow_sql ORDER BY duration_ms DESC LIMIT ?",
                    "SELECT * FROM t_slow_sql ORDER BY elapsed_ms DESC LIMIT ?",
                    "SELECT * FROM t_slow_sql ORDER BY cost_ms DESC LIMIT ?",
                    "SELECT * FROM t_slow_sql LIMIT ?",
                ],
                [max_rows],
            )
            top_sql = self._query_first_existing(
                cur,
                [
                    "SELECT * FROM top_sql ORDER BY exec_count DESC LIMIT ?",
                    "SELECT * FROM top_sql ORDER BY qps DESC LIMIT ?",
                    "SELECT * FROM top_sql ORDER BY calls DESC LIMIT ?",
                    "SELECT * FROM top_sql LIMIT ?",
                    "SELECT * FROM t_top_sql ORDER BY exec_count DESC LIMIT ?",
                    "SELECT * FROM t_top_sql ORDER BY qps DESC LIMIT ?",
                    "SELECT * FROM t_top_sql ORDER BY calls DESC LIMIT ?",
                    "SELECT * FROM t_top_sql LIMIT ?",
                ],
                [max_rows],
            )
            session_status = self._query_first_existing(
                cur,
                [
                    "SELECT * FROM session_status ORDER BY active_sessions DESC LIMIT ?",
                    "SELECT * FROM session_status ORDER BY running DESC LIMIT ?",
                    "SELECT * FROM session_status LIMIT ?",
                    "SELECT * FROM db_session_status ORDER BY active_sessions DESC LIMIT ?",
                    "SELECT * FROM db_session_status ORDER BY running DESC LIMIT ?",
                    "SELECT * FROM db_session_status LIMIT ?",
                ],
                [max_rows],
            )

            keyword_hits: List[Dict[str, Any]] = []
            lowered_keywords = [str(word or "").lower().strip() for word in (keywords or []) if str(word or "").strip()]
            if lowered_keywords:
                for row in (slow_sql + top_sql)[:200]:
                    text = json.dumps(row, ensure_ascii=False).lower()
                    if any(k and k in text for k in lowered_keywords):
                        keyword_hits.append(row)
                        if len(keyword_hits) >= max_rows:
                            break

            return {
                "engine": "sqlite",
                "db_path": str(db_path),
                "requested_tables": list(target_tables or [])[:20],
                "used_target_tables": used_target_tables,
                "fallback_reason": fallback_reason,
                "total_table_count": len(all_table_rows),
                "table_count": len(table_rows),
                "tables": table_rows[:80],
                "table_structures": table_structures[:20],
                "indexes": indexes,
                "slow_sql": slow_sql[:max_rows],
                "top_sql": top_sql[:max_rows],
                "session_status": session_status[:max_rows],
                "keyword_hits": keyword_hits[:max_rows],
            }
        finally:
            conn.close()

    @staticmethod
    def _query_first_existing(cur: sqlite3.Cursor, queries: List[str], params: List[Any]) -> List[Dict[str, Any]]:
        """执行queryfirstexisting相关逻辑，并为当前模块提供可复用的处理能力。"""
        for sql in queries:
            try:
                cur.execute(sql, params)
                return [dict(row) for row in cur.fetchall()]
            except Exception:
                continue
        return []

    @staticmethod
    async def _pg_fetch_rows(conn: Any, queries: List[str], max_rows: int) -> List[Dict[str, Any]]:
        """执行pg抓取rows相关逻辑，并为当前模块提供可复用的处理能力。"""
        for sql in queries:
            try:
                rows = await conn.fetch(sql, max_rows)
                return [dict(row) for row in rows]
            except Exception:
                continue
        return []

    @staticmethod
    def _escape_sql_identifier(name: str) -> str:
        """执行escapesqlidentifier相关逻辑，并为当前模块提供可复用的处理能力。"""
        return str(name or "").replace("'", "''")

    def _decide_tool_invocation(
        self,
        *,
        agent_name: str,
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """根据命令文本和显式开关决定本轮是否允许工具调用。"""
        return dict(decide_tool_invocation(agent_name=agent_name, assigned_command=assigned_command))

    def _command_preview(self, assigned_command: Optional[Dict[str, Any]]) -> Dict[str, Any]:
        """执行commandpreview相关逻辑，并为当前模块提供可复用的处理能力。"""
        return self._audit_builder.command_preview(assigned_command)

    def _audit(
        self,
        *,
        tool_name: str,
        action: str,
        status: str,
        detail: Dict[str, Any],
    ) -> Dict[str, Any]:
        """生成标准化工具审计记录，统一请求/响应摘要和明细预览。"""
        return self._audit_builder.build_entry(
            tool_name=tool_name,
            action=action,
            status=status,
            detail=detail,
        )

    def _next_audit_call_id(self, *, tool_name: str, action: str) -> str:
        """执行nextaudit调用id相关逻辑，并为当前模块提供可复用的处理能力。"""
        return self._audit_builder.next_call_id(tool_name=tool_name, action=action)

    def _detail_preview(self, detail: Dict[str, Any], *, max_chars: int = 420) -> str:
        """执行detailpreview相关逻辑，并为当前模块提供可复用的处理能力。"""
        try:
            text = json.dumps(detail, ensure_ascii=False, separators=(",", ":"))
        except Exception:
            text = str(detail)
        text = str(text or "").strip()
        if len(text) <= max_chars:
            return text
        return f"{text[:max_chars]}..."

    def _request_summary(self, detail: Dict[str, Any]) -> str:
        """执行request摘要相关逻辑，并为当前模块提供可复用的处理能力。"""
        picks: List[str] = []
        for key in ("path", "repo_url", "endpoint", "sheet_name", "keywords", "query", "service_name"):
            value = detail.get(key)
            if value in (None, "", [], {}):
                continue
            picks.append(f"{key}={str(value)[:100]}")
        return "；".join(picks)[:260]

    def _response_summary(self, detail: Dict[str, Any]) -> str:
        """执行response摘要相关逻辑，并为当前模块提供可复用的处理能力。"""
        picks: List[str] = []
        for key in (
            "status",
            "hits_count",
            "lines_count",
            "matches_count",
            "match_count",
            "result_count",
            "error",
        ):
            value = detail.get(key)
            if value in (None, "", [], {}):
                continue
            picks.append(f"{key}={str(value)[:100]}")
        return "；".join(picks)[:260]

    def _coerce_duration_ms(self, detail: Dict[str, Any]) -> Optional[float]:
        """执行coercedurationms相关逻辑，并为当前模块提供可复用的处理能力。"""
        for key in ("duration_ms", "latency_ms", "elapsed_ms"):
            value = detail.get(key)
            if value is None:
                continue
            try:
                return round(float(value), 2)
            except Exception:
                continue
        return None

    def _sanitize_command_part(self, item: str) -> str:
        """执行sanitizecommandpart相关逻辑，并为当前模块提供可复用的处理能力。"""
        text = str(item or "")
        masked = self._mask_url_secret(text)
        return re.sub(r"(?i)(token|apikey|api_key|access_token)=([^&\s]+)", r"\1=***", masked)

    def _mask_url_secret(self, raw_url: str) -> str:
        """执行maskurlsecret相关逻辑，并为当前模块提供可复用的处理能力。"""
        raw = str(raw_url or "").strip()
        if not raw:
            return raw
        try:
            parts = urlsplit(raw)
        except Exception:
            return raw
        if not parts.scheme or not parts.netloc:
            return raw
        netloc = parts.netloc
        if "@" in netloc:
            userinfo, host = netloc.rsplit("@", 1)
            username = userinfo.split(":", 1)[0] if userinfo else "user"
            netloc = f"{username}:***@{host}"
        safe_query = re.sub(r"(?i)(token|apikey|api_key|access_token)=([^&]+)", r"\1=***", parts.query or "")
        return urlunsplit((parts.scheme, netloc, parts.path, safe_query, parts.fragment))

    def _extract_keywords(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
    ) -> List[str]:
        """对输入执行提取keywords，将原始数据整理为稳定的内部结构。"""
        bucket: List[str] = []
        leads = self._extract_investigation_leads(compact_context, incident_context, assigned_command)
        endpoint = (((compact_context.get("interface_mapping") or {}).get("endpoint") or {}) if isinstance(compact_context.get("interface_mapping"), dict) else {})
        for key in ("path", "service", "interface", "method"):
            value = str(endpoint.get(key) or "").strip()
            if value:
                bucket.append(value)
        for field in (
            "api_endpoints",
            "service_names",
            "code_artifacts",
            "class_names",
            "monitor_items",
            "dependency_services",
            "trace_ids",
            "error_keywords",
        ):
            for value in leads.get(field) or []:
                bucket.append(value)
        for table in self._extract_database_tables(compact_context, incident_context, assigned_command):
            bucket.append(table)
        parsed = compact_context.get("parsed_data") or {}
        if isinstance(parsed, dict):
            for key in ("error_type", "error_message", "exception_class", "trace_id"):
                value = str(parsed.get(key) or "").strip()
                if value:
                    bucket.append(value)
        log_excerpt = str(compact_context.get("log_excerpt") or "")
        if log_excerpt:
            bucket.append(log_excerpt[:300])
        for key in ("task", "focus", "expected_output"):
            value = str((assigned_command or {}).get(key) or "").strip()
            if value:
                bucket.append(value)
        full_log = str(incident_context.get("log_content") or "")
        if full_log:
            bucket.append(full_log[:500])

        tokens: List[str] = []
        for raw in bucket:
            for token in re.split(r"[\s,;:|/\\\[\]\(\)\{\}\"'`]+", raw):
                tk = token.strip().lower()
                if len(tk) < 3:
                    continue
                if tk.isdigit():
                    continue
                if tk in {"http", "https", "error", "warn", "info", "debug"}:
                    continue
                tokens.append(tk[:80])
        deduped = list(dict.fromkeys(tokens))
        return deduped[:20]

    def _extract_investigation_leads(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
    ) -> Dict[str, Any]:
        """从命令、compact_context、incident_context 中抽取统一的调查线索包。"""
        picks: Dict[str, List[str]] = {
            "api_endpoints": [],
            "service_names": [],
            "code_artifacts": [],
            "class_names": [],
            "database_tables": [],
            "monitor_items": [],
            "dependency_services": [],
            "trace_ids": [],
            "error_keywords": [],
        }
        scalar = {"domain": "", "aggregate": "", "owner_team": "", "owner": ""}
        command = dict(assigned_command or {})
        sources = [
            command,
            compact_context.get("investigation_leads") if isinstance(compact_context.get("investigation_leads"), dict) else {},
            incident_context.get("investigation_leads") if isinstance(incident_context.get("investigation_leads"), dict) else {},
        ]
        for source in sources:
            if not isinstance(source, dict):
                continue
            for key in picks:
                for item in source.get(key) or []:
                    text = str(item or "").strip()
                    if text:
                        picks[key].append(text[:180])
            for key in scalar:
                if not scalar[key]:
                    scalar[key] = str(source.get(key) or "").strip()[:120]
        normalized = {key: list(dict.fromkeys(value))[:20] for key, value in picks.items()}
        normalized.update(scalar)
        return normalized

    def _extract_database_tables(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
    ) -> List[str]:
        """提取并规范化数据库表线索，统一返回有序去重后的表名列表。"""
        picks: List[str] = []
        command = dict(assigned_command or {})
        for table in command.get("database_tables") or []:
            text = str(table or "").strip()
            if text:
                picks.append(text[:120])
        interface_mapping = compact_context.get("interface_mapping")
        if isinstance(interface_mapping, dict):
            for table in (interface_mapping.get("database_tables") or interface_mapping.get("db_tables") or []):
                text = str(table or "").strip()
                if text:
                    picks.append(text[:120])
        incident_mapping = incident_context.get("interface_mapping")
        if isinstance(incident_mapping, dict):
            for table in (incident_mapping.get("database_tables") or incident_mapping.get("db_tables") or []):
                text = str(table or "").strip()
                if text:
                    picks.append(text[:120])
        for table in self._extract_investigation_leads(compact_context, incident_context, assigned_command).get("database_tables") or []:
            text = str(table or "").strip()
            if text:
                picks.append(text[:120])
        # keep order and unique
        return list(dict.fromkeys(picks))[:20]

    def _primary_service_name(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
    ) -> str:
        """返回最可信的服务名，供日志、指标、代码和变更工具共享。"""
        leads = self._extract_investigation_leads(compact_context, incident_context, assigned_command)
        candidates = [
            incident_context.get("service_name"),
            ((compact_context.get("interface_mapping") or {}).get("endpoint") or {}).get("service") if isinstance(compact_context.get("interface_mapping"), dict) else "",
            *((leads.get("service_names") or [])[:4]),
        ]
        for item in candidates:
            text = str(item or "").strip()
            if text:
                return text[:160]
        return ""

    def _primary_trace_id(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
    ) -> str:
        """返回最可信的 Trace ID，供日志/APM 等链路工具复用。"""
        leads = self._extract_investigation_leads(compact_context, incident_context, assigned_command)
        candidates = [
            incident_context.get("trace_id"),
            ((compact_context.get("parsed_data") or {}).get("trace_id") if isinstance(compact_context.get("parsed_data"), dict) else ""),
            *((leads.get("trace_ids") or [])[:4]),
        ]
        for item in candidates:
            text = str(item or "").strip()
            if text:
                return text[:160]
        return ""

    @staticmethod
    def _normalize_table_name(name: str) -> str:
        """对输入执行归一化tablename，将原始数据整理为稳定的内部结构。"""
        text = str(name or "").strip().lower().strip('"')
        if "." in text:
            return text.split(".")[-1].strip('"')
        return text

    def _search_repo(
        self,
        repo_path: str,
        keywords: List[str],
        max_hits: int,
    ) -> tuple[List[Dict[str, Any]], Dict[str, Any]]:
        """在本地代码仓执行受限搜索，返回结构化命中片段和扫描摘要。"""
        path = Path(repo_path)
        if not path.exists():
            return [], {"repo_path": repo_path, "files_scanned": 0, "hits": 0}
        hits: List[Dict[str, Any]] = []
        scanned_files = 0
        matched_files = 0
        lowered_keywords = [k.lower() for k in keywords if k]
        if not lowered_keywords:
            lowered_keywords = ["exception", "error", "timeout", "order"]

        for file in path.rglob("*"):
            if len(hits) >= max_hits:
                break
            if not file.is_file():
                continue
            if any(part in {".git", "node_modules", "dist", "build", "__pycache__"} for part in file.parts):
                continue
            if file.suffix.lower() not in SOURCE_SUFFIXES:
                continue
            if file.name.lower().startswith("test"):
                continue
            scanned_files += 1
            try:
                content = file.read_text(encoding="utf-8", errors="replace")
            except Exception:
                continue
            file_hit = False
            for index, line in enumerate(content.splitlines(), start=1):
                line_low = line.lower()
                keyword = next((kw for kw in lowered_keywords if kw in line_low), "")
                if not keyword:
                    continue
                file_hit = True
                hits.append(
                    {
                        "file": str(file.relative_to(path)),
                        "line": index,
                        "keyword": keyword,
                        "snippet": line.strip()[:220],
                    }
                )
                if len(hits) >= max_hits:
                    break
            if file_hit:
                matched_files += 1
        return hits, {
            "repo_path": str(path),
            "files_scanned": scanned_files,
            "files_with_hits": matched_files,
            "hits": len(hits),
            "keywords": lowered_keywords[:8],
        }

    def _read_log_excerpt(
        self,
        path: Path,
        max_lines: int,
        keywords: Iterable[str],
    ) -> tuple[str, int, Dict[str, Any]]:
        """从日志文件提取局部窗口，优先返回命中关键词的片段。"""
        kw = [k.lower() for k in keywords if k]
        window = deque(maxlen=max(50, max_lines))
        scanned_lines = 0
        matched_lines = 0
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            for line in handle:
                scanned_lines += 1
                text = line.rstrip("\n")
                if not kw or any(item in text.lower() for item in kw):
                    matched_lines += 1
                    window.append(text)
        if not window:
            with path.open("r", encoding="utf-8", errors="replace") as handle:
                for line in handle:
                    window.append(line.rstrip("\n"))
        lines = list(window)
        if len(lines) > max_lines:
            lines = lines[-max_lines:]
        return (
            "\n".join(lines),
            len(lines),
            {
                "file_path": str(path),
                "scanned_lines": scanned_lines,
                "matched_lines": matched_lines,
                "returned_lines": len(lines),
                "keywords": kw[:10],
            },
        )

    def _lookup_domain_file(
        self,
        path: Path,
        sheet_name: str,
        max_rows: int,
        max_matches: int,
        keywords: List[str],
    ) -> Dict[str, Any]:
        """从责任田/领域文件中查找与当前关键词最相关的记录。"""
        suffix = path.suffix.lower()
        sheet_used = ""
        if suffix == ".csv":
            rows = self._read_csv_rows(path, max_rows=max_rows)
        elif suffix in {".xlsx", ".xlsm"}:
            rows, sheet_used = self._read_xlsx_rows(path, sheet_name=sheet_name, max_rows=max_rows)
        else:
            raise RuntimeError("仅支持 .csv/.xlsx/.xlsm")

        lowered = [k.lower() for k in keywords if k]
        matches: List[Dict[str, Any]] = []
        for row in rows:
            merged = " | ".join(str(v) for v in row.values()).lower()
            if lowered and not any(k in merged for k in lowered):
                continue
            matches.append(row)
            if len(matches) >= max_matches:
                break
        if not lowered:
            matches = rows[:max_matches]
        return {
            "format": suffix,
            "sheet_used": sheet_used,
            "row_count": len(rows),
            "matches": matches,
        }

    def _read_csv_rows(self, path: Path, max_rows: int) -> List[Dict[str, Any]]:
        """负责读取csvrows，并返回后续流程可直接消费的数据结果。"""
        rows: List[Dict[str, Any]] = []
        with path.open("r", encoding="utf-8", errors="replace", newline="") as handle:
            reader = csv.DictReader(handle)
            for idx, row in enumerate(reader, start=1):
                rows.append({str(k): str(v) for k, v in (row or {}).items()})
                if idx >= max_rows:
                    break
        return rows

    def _read_xlsx_rows(self, path: Path, sheet_name: str, max_rows: int) -> tuple[List[Dict[str, Any]], str]:
        """负责读取xlsxrows，并返回后续流程可直接消费的数据结果。"""
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise RuntimeError("读取 xlsx 需要安装 openpyxl") from exc
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        headers_raw = next(rows_iter, None) or []
        headers = [str(h or f"col_{i+1}") for i, h in enumerate(headers_raw)]
        rows: List[Dict[str, Any]] = []
        for idx, row in enumerate(rows_iter, start=1):
            rows.append({headers[i]: str(value or "") for i, value in enumerate(row or []) if i < len(headers)})
            if idx >= max_rows:
                break
        wb.close()
        return rows, str(ws.title or "")

    def _resolve_log_excerpt(
        self,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        tool_context: Optional[Dict[str, Any]],
    ) -> str:
        tool_data = (tool_context or {}).get("data") if isinstance(tool_context, dict) else {}
        if isinstance(tool_data, dict):
            excerpt = str(tool_data.get("excerpt") or "").strip()
            if excerpt:
                return excerpt
        for source in (
            str(compact_context.get("log_excerpt") or "").strip(),
            str(incident_context.get("log_content") or "").strip(),
            str(incident_context.get("description") or "").strip(),
        ):
            if source:
                return source
        return ""

    def _extract_log_timeline(self, excerpt: str, *, max_events: int) -> List[Dict[str, Any]]:
        events: List[Dict[str, Any]] = []
        if not excerpt:
            return events
        for raw_line in excerpt.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            timestamp_match = re.search(r"(\d{4}-\d{2}-\d{2}[T\s]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:Z|[+-]\d{2}:?\d{2})?)", line)
            if not timestamp_match:
                timestamp_match = re.search(r"(\d{2}:\d{2}:\d{2})", line)
            level_match = re.search(r"\b(INFO|WARN|ERROR|DEBUG|TRACE)\b", line, flags=re.IGNORECASE)
            component_match = re.search(r"\b([a-zA-Z_][\w.$-]{6,})\b", line)
            events.append(
                {
                    "timestamp": str(timestamp_match.group(1) if timestamp_match else "")[:64],
                    "level": str(level_match.group(1).upper() if level_match else "")[:12],
                    "component": str(component_match.group(1) if component_match else "")[:120],
                    "message": line[:320],
                }
            )
            if len(events) >= max_events:
                break
        return events

    def _build_log_causal_timeline(self, timeline: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        chain: List[Dict[str, Any]] = []
        first_error_added = False
        resource_added = False
        user_visible_added = False
        for item in timeline:
            message = str(item.get("message") or "").lower()
            level = str(item.get("level") or "").upper()
            stage = ""
            rationale = ""
            if any(term in message for term in ("uri=", "post /api", "get /api", "request start", "createorder start")):
                stage = "request_entry"
                rationale = "请求进入系统，构成时间线起点。"
            elif not first_error_added and level == "ERROR":
                stage = "first_error"
                rationale = "首个错误事件，通常是应用侧初始异常。"
                first_error_added = True
            if not resource_added and any(term in message for term in ("connection is not available", "lock wait", "timeout after", "pending", "pool")):
                stage = "resource_exhaustion"
                rationale = "资源耗尽或阻塞放大，可能是故障扩散关键节点。"
                resource_added = True
            if any(term in message for term in ("status=502", "5xx", "upstream timeout", "bad gateway")):
                stage = "user_visible_failure"
                rationale = "用户可见错误，代表故障已经暴露到入口层。"
                user_visible_added = True
            if not stage:
                continue
            chain.append(
                {
                    "stage": stage,
                    "timestamp": str(item.get("timestamp") or ""),
                    "component": str(item.get("component") or ""),
                    "message": str(item.get("message") or "")[:280],
                    "rationale": rationale,
                }
            )
        deduped: List[Dict[str, Any]] = []
        seen = set()
        for item in chain:
            key = f"{item.get('stage')}|{item.get('timestamp')}|{item.get('component')}"
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        return deduped[:8]

    def _load_repo_focus_windows(
        self,
        *,
        repo_path: str,
        candidate_files: List[str],
        max_files: int,
        max_chars: int,
    ) -> List[Dict[str, Any]]:
        return load_repo_focus_windows(
            repo_path=repo_path,
            candidate_files=candidate_files,
            max_files=max_files,
            max_chars=max_chars,
        )

    def _expand_related_code_files(
        self,
        *,
        repo_path: str,
        seed_files: List[str],
        class_hints: List[str],
        depth: int,
        per_hop_limit: int,
    ) -> List[str]:
        return expand_related_code_files(
            repo_path=repo_path,
            seed_files=seed_files,
            class_hints=class_hints,
            depth=depth,
            per_hop_limit=per_hop_limit,
            source_suffixes=SOURCE_SUFFIXES,
        )

    def _resolve_repo_file(self, root: Path, raw_name: str) -> Optional[Path]:
        return resolve_repo_file(root, raw_name)

    def _find_symbol_file(self, root: Path, symbol: str) -> Optional[Path]:
        return find_symbol_file(root, symbol, source_suffixes=SOURCE_SUFFIXES)

    def _extract_related_code_symbols(self, text: str) -> List[str]:
        return extract_related_code_symbols(text)

    def _build_method_call_chain(
        self,
        *,
        repo_path: str,
        endpoint_interface: str,
        code_windows: List[Dict[str, Any]],
        hit_snippets: List[str],
    ) -> List[Dict[str, Any]]:
        return build_method_call_chain(
            repo_path=repo_path,
            endpoint_interface=endpoint_interface,
            code_windows=code_windows,
            hit_snippets=hit_snippets,
        )

    def _parse_interface_ref(self, raw: str) -> Dict[str, str]:
        return parse_interface_ref(raw)

    def _load_source_units(self, root: Path, files: List[str]) -> List[Dict[str, Any]]:
        return load_source_units(root, files)

    def _parse_source_unit(self, *, root: Path, file_path: Path, text: str) -> Dict[str, Any]:
        return parse_source_unit(root=root, file_path=file_path, text=text)

    def _extract_field_types(self, text: str) -> Dict[str, str]:
        return extract_field_types(text)

    def _extract_methods(self, text: str) -> Dict[str, Dict[str, Any]]:
        return extract_methods(text)

    def _guess_entry_method(self, source_units: List[Dict[str, Any]], hit_snippets: List[str]) -> str:
        return guess_entry_method(source_units, hit_snippets)

    def _find_source_unit(
        self,
        source_units: List[Dict[str, Any]],
        symbol: str,
        *,
        preferred_file: str = "",
    ) -> Optional[Dict[str, Any]]:
        return find_source_unit(source_units, symbol, preferred_file=preferred_file)

    def _resolve_next_method_call(
        self,
        *,
        source_units: List[Dict[str, Any]],
        current_unit: Dict[str, Any],
        method_meta: Dict[str, Any],
    ) -> Optional[Dict[str, str]]:
        return resolve_next_method_call(
            source_units=source_units,
            current_unit=current_unit,
            method_meta=method_meta,
        )

    @staticmethod
    def _trim_mapping(value: Any, *, item_limit: int, value_limit: int) -> Dict[str, Any]:
        if not isinstance(value, dict):
            return {}
        result: Dict[str, Any] = {}
        for key, item in list(value.items())[:item_limit]:
            if isinstance(item, list):
                result[str(key)] = item[:value_limit]
            elif isinstance(item, dict):
                result[str(key)] = dict(list(item.items())[:value_limit])
            else:
                result[str(key)] = item
        return result

    @staticmethod
    def _remote_source_summary(value: Any) -> Dict[str, Any]:
        if not isinstance(value, dict):
            return {"enabled": False, "status": "unavailable"}
        return {
            "enabled": bool(value.get("enabled")),
            "status": str(value.get("status") or ""),
            "payload_keys": list((value.get("payload") or {}).keys())[:12] if isinstance(value.get("payload"), dict) else [],
        }

    def _summarize_metric_signals(self, signals: List[Dict[str, Any]]) -> List[str]:
        summaries: List[str] = []
        for item in signals:
            metric = str(item.get("label") or item.get("metric") or "").strip()
            value = str(item.get("value") or "").strip()
            snippet = str(item.get("snippet") or "").strip()
            if not metric or not value:
                continue
            summaries.append(f"{metric}={value}，证据={snippet[:120]}")
            if len(summaries) >= 8:
                break
        return summaries

    def _build_metric_causal_chain(self, signals: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        chain: List[Dict[str, Any]] = []
        for item in signals:
            metric = str(item.get("metric") or "").strip().lower()
            label = str(item.get("label") or metric).strip()
            value = str(item.get("value") or "").strip()
            snippet = str(item.get("snippet") or "").strip()
            stage = ""
            rationale = ""
            if metric in {"cpu", "threads", "memory", "gc"}:
                stage = "resource_pressure"
                rationale = "资源指标先行异常，通常代表系统已进入压力阶段。"
            elif metric in {"db_conn", "hikari_pending", "pool", "queue"}:
                stage = "capacity_saturation"
                rationale = "连接/队列类指标打满，说明容量瓶颈已形成。"
            elif metric in {"error_rate", "latency", "p99", "availability", "5xx"}:
                stage = "user_visible_failure"
                rationale = "错误率或延迟已上升到用户可感知层。"
            if not stage:
                continue
            chain.append(
                {
                    "stage": stage,
                    "metric": metric,
                    "label": label,
                    "value": value,
                    "snippet": snippet[:180],
                    "rationale": rationale,
                }
            )
        if not chain:
            return []
        stage_order = {"resource_pressure": 1, "capacity_saturation": 2, "user_visible_failure": 3}
        chain.sort(key=lambda item: (stage_order.get(str(item.get("stage") or ""), 99), str(item.get("metric") or "")))
        deduped: List[Dict[str, Any]] = []
        seen = set()
        for item in chain:
            key = f"{item.get('stage')}|{item.get('metric')}"
            if key in seen:
                continue
            seen.add(key)
            deduped.append(item)
        return deduped[:8]

    def _extract_runbook_actions(self, items: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        actions: List[Dict[str, Any]] = []
        for item in items:
            runbook_fields = item.get("runbook_fields") if isinstance(item.get("runbook_fields"), dict) else {}
            if runbook_fields:
                steps = list(runbook_fields.get("steps") or [])[:4]
                verify = list(runbook_fields.get("verification_steps") or [])[:3]
            else:
                steps = []
                verify = []
            actions.append(
                {
                    "title": str(item.get("title") or "")[:160],
                    "entry_type": str(item.get("entry_type") or "")[:40],
                    "steps": steps,
                    "verification_steps": verify,
                }
            )
            if len(actions) >= 6:
                break
        return actions

    def _build_database_causal_summary(
        self,
        *,
        target_tables: List[str],
        tool_data: Dict[str, Any],
    ) -> Dict[str, Any]:
        slow_sql = [item for item in list(tool_data.get("slow_sql") or []) if isinstance(item, dict)]
        top_sql = [item for item in list(tool_data.get("top_sql") or []) if isinstance(item, dict)]
        session_status = [item for item in list(tool_data.get("session_status") or []) if isinstance(item, dict)]
        keyword_hits = [item for item in list(tool_data.get("keyword_hits") or []) if isinstance(item, dict)]

        evidence_points: List[str] = []
        likely_causes: List[str] = []
        dominant_pattern = "db_pressure"

        for row in session_status:
            wait_type = str(row.get("wait_event_type") or "").strip()
            wait_event = str(row.get("wait_event") or "").strip()
            sessions = row.get("sessions")
            if wait_type or wait_event:
                evidence_points.append(f"session wait: {wait_type or 'unknown'}/{wait_event or 'unknown'} sessions={sessions}")
            if wait_type.lower() == "lock" or "lock" in wait_event.lower():
                dominant_pattern = "lock_contention"
                likely_causes.append("存在数据库锁等待，事务争用可能是放大点。")

        for row in slow_sql[:3]:
            query = str(row.get("query") or row.get("sql_text") or "")[:180]
            mean_exec_time = row.get("mean_exec_time") or row.get("duration_ms") or row.get("elapsed_ms")
            if query:
                evidence_points.append(f"slow sql: {query} time={mean_exec_time}")
            if "update " in query.lower() or "for update" in query.lower():
                likely_causes.append("慢 SQL 包含写操作，可能造成锁持有时间过长。")

        for row in keyword_hits[:3]:
            query = str(row.get("query") or row.get("sql_text") or "")[:180]
            if query:
                evidence_points.append(f"keyword hit: {query}")

        if not likely_causes and slow_sql:
            likely_causes.append("数据库存在明显慢 SQL，可能导致连接占用升高和上游超时。")
        if not likely_causes and top_sql:
            likely_causes.append("数据库访问压力升高，需结合高频 SQL 判断是否存在热点查询。")

        return {
            "dominant_pattern": dominant_pattern,
            "target_tables": list(target_tables)[:12],
            "likely_causes": list(dict.fromkeys(likely_causes))[:4],
            "evidence_points": list(dict.fromkeys(evidence_points))[:6],
        }

    def _build_domain_causal_summary(
        self,
        *,
        mapping: Dict[str, Any],
        endpoint: Dict[str, Any],
        matches: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        matched = bool(mapping.get("matched"))
        owner_team = str(mapping.get("owner_team") or "")[:120]
        owner = str(mapping.get("owner") or "")[:120]
        domain = str(mapping.get("domain") or "")[:120]
        aggregate = str(mapping.get("aggregate") or "")[:120]
        feature = str(mapping.get("feature") or "")[:120]
        service = str(endpoint.get("service") or "")[:160]
        database_tables = list(mapping.get("database_tables") or mapping.get("db_tables") or [])[:12]
        dependency_services = list(mapping.get("dependency_services") or [])[:10]
        monitor_items = list(mapping.get("monitor_items") or [])[:10]

        evidence_points: List[str] = []
        if matched:
            evidence_points.append(
                f"接口已命中责任田：{domain or '-'} / {aggregate or '-'} / {feature or '-'}"
            )
        if owner_team or owner:
            evidence_points.append(f"责任团队：{owner_team or '-'}；责任人：{owner or '-'}")
        if service:
            evidence_points.append(f"接口归属服务：{service}")
        if matches:
            first = matches[0]
            evidence_points.append(
                f"责任田文档命中：{str(first.get('domain') or domain or '-')}/{str(first.get('aggregate') or aggregate or '-')}"
            )

        return {
            "dominant_pattern": "owner_confirmed" if matched and owner_team else "mapping_gap",
            "owner_team": owner_team,
            "owner": owner,
            "domain": domain,
            "aggregate": aggregate,
            "impact_scope": {
                "service": service,
                "database_tables": database_tables,
                "dependency_services": dependency_services,
                "monitor_items": monitor_items,
            },
            "evidence_points": list(dict.fromkeys(evidence_points))[:6],
        }

    def _build_change_causal_summary(
        self,
        *,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        changes: List[Dict[str, Any]],
    ) -> Dict[str, Any]:
        leads = self._extract_investigation_leads(compact_context, incident_context, assigned_command)
        endpoint_texts = [
            str(item or "").strip().lower()
            for item in list(leads.get("api_endpoints") or [])
            if str(item or "").strip()
        ]
        service_names = [
            str(item or "").strip().lower()
            for item in list(leads.get("service_names") or [])
            if str(item or "").strip()
        ]
        artifact_names = [
            str(Path(str(item or "")).stem or "").strip().lower()
            for item in list(leads.get("code_artifacts") or [])
            if str(item or "").strip()
        ]
        focus_text = " ".join(
            filter(
                None,
                [
                    str((assigned_command or {}).get("task") or "").strip(),
                    str((assigned_command or {}).get("focus") or "").strip(),
                    str(incident_context.get("description") or "").strip(),
                ],
            )
        ).lower()

        suspect_changes: List[Dict[str, Any]] = []
        mechanism_links: List[str] = []
        evidence_points: List[str] = []
        dominant_pattern = "change_window_noise"

        for change in changes[:12]:
            commit = str(change.get("commit") or "")[:12]
            subject = str(change.get("subject") or "").strip()
            subject_lower = subject.lower()
            score = 0

            if any(service and service in subject_lower for service in service_names):
                score += 2
                mechanism_links.append("变更主题直接命中责任服务，需优先评估是否为发布回归。")
            if any(endpoint and endpoint.split()[-1] in subject_lower for endpoint in endpoint_texts):
                score += 2
                mechanism_links.append("变更主题与问题接口路径相邻，可能影响路由或接口注册。")
            if any(artifact and artifact in subject_lower for artifact in artifact_names):
                score += 2
                mechanism_links.append("变更主题提到责任田代码符号，可能直接触发实现回归。")
            if any(token in subject_lower for token in ("route", "mapping", "controller", "retry", "timeout", "transaction", "pool")):
                score += 1
                mechanism_links.append("变更涉及路由/重试/事务/连接池等敏感机制，可能改变故障放大链。")
            if focus_text and any(token in subject_lower for token in focus_text.split()):
                score += 1

            if subject:
                evidence_points.append(f"{commit}: {subject[:160]}")
            if score >= 2:
                dominant_pattern = "recent_release_regression"
                suspect_changes.append(
                    {
                        "commit": commit,
                        "author": str(change.get("author") or "")[:80],
                        "time": str(change.get("time") or "")[:64],
                        "subject": subject[:200],
                        "score": score,
                    }
                )

        if dominant_pattern == "recent_release_regression":
            mechanism_links.append("近期变更与服务/接口/关键机制同时命中，符合发布后回归的初步特征。")
        elif changes:
            mechanism_links.append("已获取变更窗口，但暂无强匹配项，需要与日志和代码证据交叉验证。")
        else:
            evidence_points.append("变更窗口为空，当前无法从代码提交侧建立直接因果。")

        return {
            "dominant_pattern": dominant_pattern,
            "suspect_changes": suspect_changes[:4],
            "mechanism_links": list(dict.fromkeys(mechanism_links))[:4],
            "evidence_points": list(dict.fromkeys(evidence_points))[:6],
        }

    def _build_runbook_action_summary(
        self,
        *,
        compact_context: Dict[str, Any],
        incident_context: Dict[str, Any],
        assigned_command: Optional[Dict[str, Any]],
        items: List[Dict[str, Any]],
        recommended_actions: List[Dict[str, Any]],
        source: str,
    ) -> Dict[str, Any]:
        service_name = self._primary_service_name(compact_context, incident_context, assigned_command)
        recommended_steps: List[str] = []
        verification_steps: List[str] = []
        evidence_points: List[str] = []
        dominant_pattern = "knowledge_gap"

        for item in items[:6]:
            title = str(item.get("title") or "").strip()
            entry_type = str(item.get("entry_type") or "").strip().lower()
            summary = str(item.get("summary") or "").strip()
            runbook_fields = item.get("runbook_fields") if isinstance(item.get("runbook_fields"), dict) else {}
            steps = [str(value).strip() for value in list(runbook_fields.get("steps") or []) if str(value).strip()]
            verify = [
                str(value).strip()
                for value in list(runbook_fields.get("verification_steps") or [])
                if str(value).strip()
            ]
            if title:
                evidence_points.append(f"{entry_type or 'knowledge'}: {title[:160]}")
            if summary:
                evidence_points.append(summary[:180])
            if entry_type == "runbook" and steps:
                dominant_pattern = "matched_runbook"
                recommended_steps.extend(steps[:3])
                verification_steps.extend(verify[:3])

        if dominant_pattern == "knowledge_gap" and recommended_actions:
            for action in recommended_actions[:3]:
                recommended_steps.extend(
                    [str(value).strip() for value in list(action.get("steps") or []) if str(value).strip()][:2]
                )
                verification_steps.extend(
                    [
                        str(value).strip()
                        for value in list(action.get("verification_steps") or [])
                        if str(value).strip()
                    ][:2]
                )
            if recommended_steps:
                dominant_pattern = "matched_runbook"

        if dominant_pattern == "matched_runbook":
            evidence_points.append(
                f"已命中 {source or 'knowledge'} 中与 {service_name or '当前服务'} 相关的处置知识，可用于止血与验证。"
            )
        else:
            evidence_points.append("当前未命中可直接执行的 Runbook，需要结合专家结论补充处置步骤。")

        return {
            "dominant_pattern": dominant_pattern,
            "service_name": service_name,
            "recommended_steps": list(dict.fromkeys(recommended_steps))[:5],
            "verification_steps": list(dict.fromkeys(verification_steps))[:4],
            "evidence_points": list(dict.fromkeys(evidence_points))[:6],
        }


agent_tool_context_service = AgentToolContextService()
