"""
三态资产服务模块

本模块提供三态资产的统一管理功能，包括：
1. 运行态资产（RuntimeAsset）：日志、指标、追踪数据
2. 开发态资产（DevAsset）：代码、配置、文档
3. 设计态资产（DesignAsset）：领域模型、架构文档、案例库
4. 责任田资产：接口到领域、聚合根的映射关系

资产关联设计：
- 运行态资产（故障现象）-> 开发态资产（代码实现）-> 设计态资产（领域设计）
- 通过 link_assets 建立关联链路，支持根因追踪

责任田资产：
- 用户可导入/维护的责任田映射表
- 支持从 Excel/CSV 批量导入
- 用于接口上下文定位（DomainAgent 使用）

Tri-State Asset Service
"""

import csv
from io import BytesIO, StringIO
import json
from pathlib import Path
import re
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

import structlog

from app.config import settings
from app.models.asset import (
    TriStateAsset,
    RuntimeAsset,
    DevAsset,
    DesignAsset,
    RuntimeAssetType,
    DevAssetType,
    DesignAssetType,
    DomainModel,
    CaseLibrary,
)
from app.repositories.asset_repository import (
    AssetRepository,
    InMemoryAssetRepository,
)
from app.services.asset_knowledge_service import asset_knowledge_service

logger = structlog.get_logger()


class AssetService:
    """
    三态资产服务

    管理三类资产的完整生命周期：
    - RuntimeAsset: 运行态资产（日志、指标、追踪）
    - DevAsset: 开发态资产（代码、配置、文档）
    - DesignAsset: 设计态资产（领域模型、架构文档）

    核心功能：
    1. 资产 CRUD 操作
    2. 资产关联追踪
    3. 责任田资产管理
    4. 接口上下文定位
    """
    
    def __init__(self, repository: Optional[AssetRepository] = None):
        """
        初始化资产服务

        Args:
            repository: 资产存储库，未提供则使用内存存储
        """
        self._repository = repository or InMemoryAssetRepository()
        self._sample_bootstrapped = False
        self._bootstrap_repo_id = id(self._repository)
        # 责任田资产存储路径
        store_root = Path(settings.LOCAL_STORE_DIR) / "assets"
        store_root.mkdir(parents=True, exist_ok=True)
        self._responsibility_asset_file = store_root / "responsibility_assets.json"

    async def _ensure_sample_knowledge_loaded(self) -> None:
        """
        确保示例知识库已加载

        将本地 Markdown 示例注入到内存仓储。
        仅在首次调用或仓储实例更换后执行一次。
        """
        if id(self._repository) != self._bootstrap_repo_id:
            self._sample_bootstrapped = False
            self._bootstrap_repo_id = id(self._repository)

        if self._sample_bootstrapped:
            return

        # 加载领域模型和案例
        payload = asset_knowledge_service.build_bootstrap_models()
        domain_models = payload.get("domain_models", [])
        cases = payload.get("cases", [])

        for model in domain_models:
            if not await self._repository.get_domain_model(model.name):
                await self._repository.save_domain_model(model)

        for case in cases:
            if not await self._repository.get_case(case.id):
                await self._repository.save_case(case)

        self._sample_bootstrapped = True

    # ==============================================================================
    # 运行态资产（RuntimeAsset）
    # ==============================================================================
    
    async def create_runtime_asset(
        self,
        type: RuntimeAssetType,
        source: str,
        raw_content: Optional[str] = None,
        parsed_data: Optional[Dict[str, Any]] = None,
        service_name: Optional[str] = None,
        trace_id: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> RuntimeAsset:
        """
        创建运行态资产

        运行态资产代表生产环境中的实时数据：
        - LOG: 应用日志
        - METRIC: 监控指标
        - TRACE: 分布式追踪
        - EVENT: 业务事件

        Args:
            type: 资产类型
            source: 数据来源
            raw_content: 原始内容
            parsed_data: 解析后的结构化数据
            service_name: 服务名称
            trace_id: 追踪 ID
            metadata: 元数据

        Returns:
            RuntimeAsset: 创建的运行态资产
        """
        asset_id = f"rt_{uuid.uuid4().hex[:8]}"

        asset = RuntimeAsset(
            id=asset_id,
            type=type,
            source=source,
            raw_content=raw_content,
            parsed_data=parsed_data,
            service_name=service_name,
            trace_id=trace_id,
            metadata=metadata or {}
        )

        await self._repository.save_runtime_asset(asset)

        logger.info(
            "runtime_asset_created",
            asset_id=asset_id,
            type=type,
            source=source
        )

        return asset

    async def get_runtime_asset(self, asset_id: str) -> Optional[RuntimeAsset]:
        """获取运行态资产"""
        return await self._repository.get_runtime_asset(asset_id)

    async def list_runtime_assets(
        self,
        type: Optional[RuntimeAssetType] = None,
        service_name: Optional[str] = None
    ) -> List[RuntimeAsset]:
        """列出运行态资产，支持按类型和服务名过滤"""
        assets = await self._repository.list_runtime_assets()

        if type:
            assets = [a for a in assets if a.type == type]
        if service_name:
            assets = [a for a in assets if a.service_name == service_name]

        return assets

    # ==============================================================================
    # 开发态资产（DevAsset）
    # ==============================================================================
    
    async def create_dev_asset(
        self,
        type: DevAssetType,
        name: str,
        path: str,
        language: Optional[str] = None,
        content: Optional[str] = None,
        repo_url: Optional[str] = None,
        branch: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> DevAsset:
        """
        创建开发态资产

        开发态资产代表代码仓库中的静态资源：
        - CODE: 源代码文件
        - CONFIG: 配置文件
        - DOC: 文档文件
        - TEST: 测试文件

        Args:
            type: 资产类型
            name: 资产名称
            path: 文件路径
            language: 编程语言
            content: 文件内容
            repo_url: 仓库 URL
            branch: 分支名
            metadata: 元数据

        Returns:
            DevAsset: 创建的开发态资产
        """
        asset_id = f"dev_{uuid.uuid4().hex[:8]}"

        asset = DevAsset(
            id=asset_id,
            type=type,
            name=name,
            path=path,
            language=language,
            content=content,
            repo_url=repo_url,
            branch=branch,
            metadata=metadata or {}
        )

        await self._repository.save_dev_asset(asset)

        logger.info(
            "dev_asset_created",
            asset_id=asset_id,
            type=type,
            name=name
        )

        return asset

    async def get_dev_asset(self, asset_id: str) -> Optional[DevAsset]:
        """获取开发态资产"""
        return await self._repository.get_dev_asset(asset_id)

    async def list_dev_assets(
        self,
        type: Optional[DevAssetType] = None,
        language: Optional[str] = None
    ) -> List[DevAsset]:
        """列出开发态资产，支持按类型和语言过滤"""
        assets = await self._repository.list_dev_assets()

        if type:
            assets = [a for a in assets if a.type == type]
        if language:
            assets = [a for a in assets if a.language == language]

        return assets

    async def search_code(self, query: str) -> List[DevAsset]:
        """
        搜索代码资产

        在代码资产的名称、内容和解析数据中搜索关键词。

        Args:
            query: 搜索关键词（类名、方法名等）

        Returns:
            List[DevAsset]: 匹配的代码资产列表
        """
        results = []
        query_lower = query.lower()

        for asset in await self._repository.list_dev_assets():
            if asset.type != DevAssetType.CODE:
                continue

            # 搜索名称
            if query_lower in asset.name.lower():
                results.append(asset)
                continue

            # 搜索内容
            if asset.content and query_lower in asset.content.lower():
                results.append(asset)
                continue

            # 搜索解析数据
            if asset.parsed_data:
                parsed_str = str(asset.parsed_data).lower()
                if query_lower in parsed_str:
                    results.append(asset)

        return results

    # ==============================================================================
    # 设计态资产（DesignAsset）
    # ==============================================================================
    
    async def create_design_asset(
        self,
        type: DesignAssetType,
        name: str,
        content: Optional[str] = None,
        parsed_data: Optional[Dict[str, Any]] = None,
        domain: Optional[str] = None,
        owner: Optional[str] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> DesignAsset:
        """
        创建设计态资产

        设计态资产代表架构设计文档：
        - ARCHITECTURE: 架构文档
        - API_SPEC: API 规范
        - DOMAIN_MODEL: 领域模型
        - CASE_LIBRARY: 案例库

        Args:
            type: 资产类型
            name: 资产名称
            content: 文档内容
            parsed_data: 解析后的结构化数据
            domain: 所属领域
            owner: 负责人
            metadata: 元数据

        Returns:
            DesignAsset: 创建的设计态资产
        """
        asset_id = f"des_{uuid.uuid4().hex[:8]}"

        asset = DesignAsset(
            id=asset_id,
            type=type,
            name=name,
            content=content,
            parsed_data=parsed_data,
            domain=domain,
            owner=owner,
            metadata=metadata or {}
        )

        await self._repository.save_design_asset(asset)

        logger.info(
            "design_asset_created",
            asset_id=asset_id,
            type=type,
            name=name
        )

        return asset

    async def get_design_asset(self, asset_id: str) -> Optional[DesignAsset]:
        """获取设计态资产"""
        return await self._repository.get_design_asset(asset_id)

    async def list_design_assets(
        self,
        type: Optional[DesignAssetType] = None,
        domain: Optional[str] = None
    ) -> List[DesignAsset]:
        """列出设计态资产，支持按类型和领域过滤"""
        assets = await self._repository.list_design_assets()

        if type:
            assets = [a for a in assets if a.type == type]
        if domain:
            assets = [a for a in assets if a.domain == domain]

        return assets

    # ==============================================================================
    # 领域模型（DomainModel）
    # ==============================================================================
    
    async def create_domain_model(
        self,
        name: str,
        description: Optional[str] = None,
        aggregates: Optional[List[str]] = None,
        entities: Optional[List[str]] = None,
        owner_team: Optional[str] = None
    ) -> DomainModel:
        """
        创建领域模型

        领域模型定义了 DDD 中的领域概念：
        - name: 领域名称（如"订单域"）
        - aggregates: 聚合根列表
        - entities: 实体列表
        - owner_team: 责任团队

        Args:
            name: 领域名称
            description: 描述
            aggregates: 聚合根列表
            entities: 实体列表
            owner_team: 责任团队

        Returns:
            DomainModel: 创建的领域模型
        """
        model = DomainModel(
            name=name,
            description=description,
            aggregates=aggregates or [],
            entities=entities or [],
            owner_team=owner_team
        )

        await self._repository.save_domain_model(model)

        logger.info(
            "domain_model_created",
            name=name,
            aggregates=len(aggregates or [])
        )

        return model

    async def get_domain_model(self, name: str) -> Optional[DomainModel]:
        """获取领域模型"""
        await self._ensure_sample_knowledge_loaded()
        return await self._repository.get_domain_model(name)

    async def list_domain_models(self) -> List[DomainModel]:
        """列出所有领域模型"""
        await self._ensure_sample_knowledge_loaded()
        return await self._repository.list_domain_models()

    async def find_domain_by_aggregate(self, aggregate_name: str) -> Optional[DomainModel]:
        """
        根据聚合名称查找领域

        用于 DomainAgent 定位接口所属的领域。

        Args:
            aggregate_name: 聚合名称

        Returns:
            Optional[DomainModel]: 领域模型或 None
        """
        await self._ensure_sample_knowledge_loaded()
        for model in await self._repository.list_domain_models():
            if aggregate_name in model.aggregates:
                return model
        return None

    # ==============================================================================
    # 案例库（CaseLibrary）
    # ==============================================================================
    
    async def create_case(
        self,
        title: str,
        description: str,
        incident_type: str,
        root_cause: str,
        solution: str,
        symptoms: Optional[List[str]] = None,
        related_services: Optional[List[str]] = None,
        tags: Optional[List[str]] = None
    ) -> CaseLibrary:
        """
        创建案例

        案例库存储历史故障处理经验，用于：
        - 相似案例检索
        - 处置建议生成
        - 知识沉淀

        Args:
            title: 案例标题
            description: 案例描述
            incident_type: 故障类型
            root_cause: 根因分析
            solution: 解决方案
            symptoms: 症状列表
            related_services: 相关服务
            tags: 标签

        Returns:
            CaseLibrary: 创建的案例
        """
        case_id = f"case_{uuid.uuid4().hex[:8]}"

        case = CaseLibrary(
            id=case_id,
            title=title,
            description=description,
            incident_type=incident_type,
            symptoms=symptoms or [],
            root_cause=root_cause,
            root_cause_category=incident_type,
            solution=solution,
            fix_steps=[],
            related_services=related_services or [],
            related_code=[],
            tags=tags or []
        )

        await self._repository.save_case(case)

        logger.info(
            "case_created",
            case_id=case_id,
            title=title,
            incident_type=incident_type
        )

        return case

    async def get_case(self, case_id: str) -> Optional[CaseLibrary]:
        """获取案例"""
        await self._ensure_sample_knowledge_loaded()
        return await self._repository.get_case(case_id)

    async def list_cases(
        self,
        incident_type: Optional[str] = None,
        tag: Optional[str] = None
    ) -> List[CaseLibrary]:
        """列出案例，支持按类型和标签过滤"""
        await self._ensure_sample_knowledge_loaded()
        cases = await self._repository.list_cases()

        if incident_type:
            cases = [c for c in cases if c.incident_type == incident_type]
        if tag:
            cases = [c for c in cases if tag in c.tags]

        return cases

    async def search_similar_cases(
        self,
        symptoms: List[str],
        exception_type: Optional[str] = None,
        limit: int = 5
    ) -> List[CaseLibrary]:
        """
        搜索相似案例

        根据症状和异常类型匹配历史案例。
        匹配规则：
        - 异常类型匹配：+10 分
        - 症状在描述中：+1 分
        - 症状在症状列表中：+2 分

        Args:
            symptoms: 症状列表
            exception_type: 异常类型
            limit: 返回数量限制

        Returns:
            List[CaseLibrary]: 相似案例列表（按分数排序）
        """
        await self._ensure_sample_knowledge_loaded()
        results = []

        for case in await self._repository.list_cases():
            score = 0

            # 匹配异常类型
            if exception_type and exception_type in case.root_cause:
                score += 10

            # 匹配症状
            for symptom in symptoms:
                if symptom.lower() in case.description.lower():
                    score += 1
                for case_symptom in case.symptoms:
                    if symptom.lower() in case_symptom.lower():
                        score += 2

            if score > 0:
                results.append((case, score))

        # 按分数排序
        results.sort(key=lambda x: x[1], reverse=True)

        return [r[0] for r in results[:limit]]

    async def locate_interface_context(
        self,
        log_content: str,
        symptom: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        根据日志/现象中的接口信息，定位领域-聚合根责任田。
        """
        await self._ensure_sample_knowledge_loaded()
        local_hit = await self._locate_from_responsibility_assets(log_content=log_content, symptom=symptom)
        if local_hit:
            return local_hit
        return asset_knowledge_service.locate_by_log(
            log_content=log_content or "",
            symptom=symptom,
        )

    async def locate_responsibility_assets(
        self,
        *,
        log_content: str,
        symptom: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """仅使用系统内维护的责任田资产进行匹配。"""
        await self._ensure_sample_knowledge_loaded()
        return await self._locate_from_responsibility_assets(
            log_content=log_content,
            symptom=symptom,
        )

    # ============== 责任田资产（用户维护） ==============

    async def list_responsibility_assets(
        self,
        *,
        query: Optional[str] = None,
        domain: Optional[str] = None,
        aggregate: Optional[str] = None,
        api_keyword: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """查询用户维护的责任田资产，并支持全文、领域、聚合根、接口关键词过滤。"""
        rows = self._load_responsibility_assets()
        q = str(query or "").strip().lower()
        domain_q = str(domain or "").strip().lower()
        aggregate_q = str(aggregate or "").strip().lower()
        api_q = str(api_keyword or "").strip().lower()

        def _hit(row: Dict[str, Any]) -> bool:
            """判断单条责任田记录是否命中过滤条件。"""
            if q:
                corpus = json.dumps(row, ensure_ascii=False).lower()
                if q not in corpus:
                    return False
            if domain_q and domain_q not in str(row.get("domain") or "").lower():
                return False
            if aggregate_q and aggregate_q not in str(row.get("aggregate") or "").lower():
                return False
            if api_q:
                api_text = " ".join(str(x) for x in list(row.get("api_interfaces") or []))
                if api_q not in api_text.lower():
                    return False
            return True

        filtered = [row for row in rows if _hit(row)]
        filtered.sort(key=lambda x: str(x.get("updated_at") or ""), reverse=True)
        return filtered

    async def upsert_responsibility_asset(self, payload: Dict[str, Any]) -> Dict[str, Any]:
        """新增或更新一条责任田资产记录。"""
        rows = self._load_responsibility_assets()
        now = datetime.utcnow().isoformat()
        row = self._normalize_responsibility_row(payload, source_file="manual", row_index=None)
        existing_idx = -1
        asset_id = str(row.get("asset_id") or "").strip()
        if asset_id:
            for idx, item in enumerate(rows):
                if str(item.get("asset_id") or "").strip() == asset_id:
                    existing_idx = idx
                    break
        if existing_idx >= 0:
            row["created_at"] = rows[existing_idx].get("created_at") or now
            row["updated_at"] = now
            rows[existing_idx] = row
        else:
            if not asset_id:
                row["asset_id"] = f"own_{uuid.uuid4().hex[:12]}"
            row["created_at"] = row.get("created_at") or now
            row["updated_at"] = now
            rows.append(row)
        self._save_responsibility_assets(rows)
        return row

    async def delete_responsibility_asset(self, asset_id: str) -> bool:
        """删除指定责任田资产记录。"""
        key = str(asset_id or "").strip()
        if not key:
            return False
        rows = self._load_responsibility_assets()
        next_rows = [row for row in rows if str(row.get("asset_id") or "").strip() != key]
        if len(next_rows) == len(rows):
            return False
        self._save_responsibility_assets(next_rows)
        return True

    async def import_responsibility_assets_from_file(
        self,
        *,
        file_name: str,
        file_bytes: bytes,
        replace_existing: bool = True,
    ) -> Dict[str, Any]:
        """从 Excel/CSV 批量导入责任田资产，并支持替换或合并。"""
        rows_from_file = self._parse_responsibility_file(file_name=file_name, file_bytes=file_bytes)
        normalized: List[Dict[str, Any]] = []
        for idx, raw in enumerate(rows_from_file, start=1):
            try:
                normalized.append(
                    self._normalize_responsibility_row(raw, source_file=file_name, row_index=idx)
                )
            except Exception as exc:
                logger.warning(
                    "responsibility_asset_row_invalid",
                    file_name=file_name,
                    row_index=idx,
                    error=str(exc),
                )

        if replace_existing:
            merged = normalized
        else:
            existing = self._load_responsibility_assets()
            by_key: Dict[str, Dict[str, Any]] = {}
            for row in existing:
                key = self._responsibility_row_key(row)
                by_key[key] = row
            for row in normalized:
                key = self._responsibility_row_key(row)
                prev = by_key.get(key)
                if prev and not row.get("asset_id"):
                    row["asset_id"] = prev.get("asset_id")
                    row["created_at"] = prev.get("created_at")
                by_key[key] = row
            merged = list(by_key.values())
        now = datetime.utcnow().isoformat()
        for row in merged:
            if not str(row.get("asset_id") or "").strip():
                row["asset_id"] = f"own_{uuid.uuid4().hex[:12]}"
            row["created_at"] = row.get("created_at") or now
            row["updated_at"] = now
        self._save_responsibility_assets(merged)

        return {
            "file_name": file_name,
            "replace_existing": bool(replace_existing),
            "imported": len(normalized),
            "stored": len(merged),
            "preview": merged[:20],
        }

    def responsibility_asset_schema(self) -> Dict[str, Any]:
        """返回责任田资产模板字段、别名映射和导入提示。"""
        return {
            "required_fields": [
                "feature",
                "domain",
                "aggregate",
                "frontend_pages",
                "api_interfaces",
                "code_items",
                "database_tables",
                "dependency_services",
                "monitor_items",
            ],
            "aliases": self._responsibility_header_aliases(),
            "tips": [
                "列表字段支持使用逗号、分号、顿号、换行分隔",
                "api_interfaces 建议写为：POST /api/v1/orders",
            ],
        }

    async def responsibility_asset_stats(self) -> Dict[str, Any]:
        """返回责任田资产统计信息和实际存储路径。"""
        rows = self._load_responsibility_assets()
        return {
            "count": len(rows),
            "latest_updated_at": max((str(x.get("updated_at") or "") for x in rows), default=""),
            "storage_path": str(self._responsibility_asset_file),
        }

    async def _locate_from_responsibility_assets(
        self,
        *,
        log_content: str,
        symptom: Optional[str],
    ) -> Optional[Dict[str, Any]]:
        """优先用用户维护的责任田资产匹配接口上下文，命中后直接返回结构化线索。"""
        rows = self._load_responsibility_assets()
        if not rows:
            return None
        corpus = "\n".join(x for x in [str(log_content or ""), str(symptom or "")] if x).strip()
        if not corpus:
            return None
        hints = self._extract_interface_hints(corpus)
        if not hints:
            return None
        best_score = -1
        best_row: Optional[Dict[str, Any]] = None
        best_endpoint: Optional[Dict[str, Any]] = None
        for row in rows:
            for endpoint in list(row.get("api_interfaces") or []):
                method, path = self._parse_method_path(endpoint)
                path_norm = self._normalize_path(path)
                if not path_norm:
                    continue
                regex = self._path_template_to_regex(path_norm)
                for hint in hints:
                    hint_method = str(hint.get("method") or "").upper()
                    hint_path = self._normalize_path(str(hint.get("path") or ""))
                    if not hint_path:
                        continue
                    score = 0
                    if re.match(regex, hint_path):
                        score += 8
                        if hint_path == path_norm and "{" not in path_norm:
                            score += 2
                    elif hint_path.endswith(path_norm) or path_norm.endswith(hint_path):
                        score += 5
                    if hint_method and method and hint_method == method:
                        score += 3
                    if score > best_score:
                        best_score = score
                        best_row = row
                        best_endpoint = {
                            "method": method,
                            "path": path_norm,
                            "service": ",".join(list(row.get("dependency_services") or [])[:2]),
                            "interface": endpoint,
                            "matched_hint": hint,
                        }
        if best_score < 4 or not best_row:
            return None
        code_items = [str(x).strip() for x in list(best_row.get("code_items") or []) if str(x).strip()]
        code_artifacts = [{"path": item, "symbol": item} for item in code_items[:20]]
        db_tables = [str(x).strip() for x in list(best_row.get("database_tables") or []) if str(x).strip()]
        confidence = min(0.99, 0.42 + best_score * 0.05)
        return {
            "matched": True,
            "confidence": round(confidence, 3),
            "source": "responsibility_assets",
            "reason": "已命中用户维护的责任田资产映射",
            "guidance": [],
            "interface_hints": hints[:10],
            "domain": str(best_row.get("domain") or ""),
            "aggregate": str(best_row.get("aggregate") or ""),
            "owner_team": str(best_row.get("owner_team") or ""),
            "owner": str(best_row.get("owner") or ""),
            "responsibility_asset_id": str(best_row.get("asset_id") or ""),
            "matched_endpoint": best_endpoint,
            "code_artifacts": code_artifacts,
            "db_tables": db_tables,
            "database_tables": list(db_tables),
            "dependency_services": list(best_row.get("dependency_services") or []),
            "monitor_items": list(best_row.get("monitor_items") or []),
            "design_ref": {
                "doc": "责任田资产",
                "section": str(best_row.get("feature") or ""),
            },
            "design_details": {
                "description": f"特性：{best_row.get('feature') or '-'}",
                "entities": [],
                "value_objects": [],
                "domain_services": list(best_row.get("dependency_services") or []),
                "invariants": [],
                "events": [],
            },
            "similar_cases": [],
        }

    def _parse_responsibility_file(self, *, file_name: str, file_bytes: bytes) -> List[Dict[str, Any]]:
        """解析责任田导入文件，兼容 Excel 与 CSV 文本。"""
        suffix = Path(file_name).suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
            wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(x or "").strip() for x in rows[0]]
            result: List[Dict[str, Any]] = []
            for row in rows[1:]:
                if row is None:
                    continue
                payload = {
                    headers[idx]: row[idx] if idx < len(row) else None
                    for idx in range(len(headers))
                    if str(headers[idx] or "").strip()
                }
                if any(str(v or "").strip() for v in payload.values()):
                    result.append(payload)
            return result
        text = None
        for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
            try:
                text = file_bytes.decode(encoding)
                break
            except Exception:
                continue
        if text is None:
            raise ValueError("无法解析文件编码，仅支持 UTF-8/GBK")
        reader = csv.DictReader(StringIO(text))
        return [dict(row or {}) for row in reader]

    def _responsibility_header_aliases(self) -> Dict[str, List[str]]:
        """定义责任田模板字段与常见表头别名映射。"""
        return {
            "feature": ["特性", "feature", "业务特性"],
            "domain": ["领域", "domain"],
            "aggregate": ["聚合根", "aggregate", "aggregate_root"],
            "frontend_pages": ["前端页面", "frontend_pages", "pages", "页面"],
            "api_interfaces": ["api接口", "api", "api_interfaces", "接口"],
            "code_items": ["代码清单", "code_items", "code", "代码"],
            "database_tables": ["数据库表", "database_tables", "db_tables", "table", "tables"],
            "dependency_services": ["依赖服务", "dependency_services", "dependencies", "下游服务"],
            "monitor_items": ["监控清单", "monitor_items", "monitors", "监控项"],
            "owner_team": ["责任团队", "owner_team", "团队"],
            "owner": ["负责人", "owner"],
        }

    def _normalize_responsibility_row(
        self,
        payload: Dict[str, Any],
        *,
        source_file: str,
        row_index: Optional[int],
    ) -> Dict[str, Any]:
        """把导入行或手工输入规范化为统一责任田记录结构。"""
        alias = self._responsibility_header_aliases()
        row: Dict[str, Any] = {}
        for key, candidates in alias.items():
            value = None
            for name in candidates:
                if name in payload and payload.get(name) is not None:
                    value = payload.get(name)
                    break
            row[key] = value
        now = datetime.utcnow().isoformat()
        normalized = {
            "asset_id": str(payload.get("asset_id") or "").strip(),
            "feature": self._norm_text(row.get("feature")),
            "domain": self._norm_text(row.get("domain")),
            "aggregate": self._norm_text(row.get("aggregate")),
            "frontend_pages": self._split_list_field(row.get("frontend_pages")),
            "api_interfaces": self._split_list_field(row.get("api_interfaces")),
            "code_items": self._split_list_field(row.get("code_items")),
            "database_tables": self._split_list_field(row.get("database_tables")),
            "dependency_services": self._split_list_field(row.get("dependency_services")),
            "monitor_items": self._split_list_field(row.get("monitor_items")),
            "owner_team": self._norm_text(row.get("owner_team")),
            "owner": self._norm_text(row.get("owner")),
            "source_file": source_file,
            "row_index": row_index,
            "created_at": str(payload.get("created_at") or now),
            "updated_at": str(payload.get("updated_at") or now),
        }
        if not normalized["domain"] or not normalized["aggregate"]:
            raise ValueError("domain/aggregate 必填")
        if not normalized["api_interfaces"]:
            raise ValueError("api_interfaces 至少包含一项")
        if not normalized["asset_id"]:
            normalized["asset_id"] = f"own_{uuid.uuid4().hex[:12]}"
        return normalized

    def _responsibility_row_key(self, row: Dict[str, Any]) -> str:
        """生成责任田记录的自然键，用于导入去重合并。"""
        return "|".join(
            [
                str(row.get("feature") or "").strip().lower(),
                str(row.get("domain") or "").strip().lower(),
                str(row.get("aggregate") or "").strip().lower(),
                ",".join(sorted(str(x).strip().lower() for x in list(row.get("api_interfaces") or []))),
            ]
        )

    def _load_responsibility_assets(self) -> List[Dict[str, Any]]:
        """从本地 JSON 文件读取责任田资产列表。"""
        if not self._responsibility_asset_file.exists():
            return []
        try:
            payload = json.loads(self._responsibility_asset_file.read_text(encoding="utf-8"))
            if isinstance(payload, list):
                return [dict(x) for x in payload if isinstance(x, dict)]
            return []
        except Exception:
            return []

    def _save_responsibility_assets(self, rows: List[Dict[str, Any]]) -> None:
        """把责任田资产列表原子写回本地文件。"""
        tmp = self._responsibility_asset_file.with_suffix(".json.tmp")
        tmp.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(self._responsibility_asset_file)

    def _split_list_field(self, value: Any) -> List[str]:
        """把逗号、分号、换行等分隔的列表字段规范化为去重字符串列表。"""
        if value is None:
            return []
        if isinstance(value, list):
            items = value
        else:
            text = str(value).strip()
            if not text:
                return []
            items = re.split(r"[,\n;；、|]+", text)
        output: List[str] = []
        seen = set()
        for item in items:
            val = self._norm_text(item)
            if val and val not in seen:
                seen.add(val)
                output.append(val)
        return output

    def _norm_text(self, value: Any) -> str:
        """对任意值做安全字符串化并去首尾空白。"""
        if value is None:
            return ""
        return str(value).strip()

    def _extract_interface_hints(self, text: str) -> List[Dict[str, str]]:
        """从日志或症状文本中提取 HTTP 方法和接口路径提示。"""
        hints: List[Dict[str, str]] = []
        seen = set()
        pattern = re.compile(
            r"\b(GET|POST|PUT|PATCH|DELETE)\s+((?:https?://[^\s\"']+)?/[A-Za-z0-9_\-./{}]+)",
            flags=re.IGNORECASE,
        )
        for method, raw in pattern.findall(text):
            path = self._normalize_path(raw)
            key = f"{method.upper()} {path}"
            if path and key not in seen:
                seen.add(key)
                hints.append({"method": method.upper(), "path": path})
        generic = re.compile(r"(/(?:[A-Za-z0-9_\-{}]+)(?:/[A-Za-z0-9_\-{}]+)*)")
        for raw in generic.findall(text):
            path = self._normalize_path(raw)
            key = f"ANY {path}"
            if path and key not in seen:
                seen.add(key)
                hints.append({"method": "", "path": path})
        return hints[:12]

    def _parse_method_path(self, endpoint: str) -> tuple[str, str]:
        """把 `METHOD /path` 形式的接口字符串拆成方法和路径。"""
        text = str(endpoint or "").strip()
        match = re.match(r"^(GET|POST|PUT|PATCH|DELETE)\s+(.+)$", text, flags=re.IGNORECASE)
        if not match:
            return "", text
        return match.group(1).upper(), match.group(2).strip()

    def _normalize_path(self, raw_path: str) -> str:
        """规范化 URL 路径，移除域名、查询串和尾随分隔符。"""
        path = str(raw_path or "").strip().rstrip('.,;:)')
        if "//" in path and path.startswith("http"):
            match = re.match(r"https?://[^/]+(/.*)", path)
            if match:
                path = match.group(1)
        if "?" in path:
            path = path.split("?", 1)[0]
        if "#" in path:
            path = path.split("#", 1)[0]
        if path and not path.startswith("/"):
            path = "/" + path
        if len(path) > 1 and path.endswith("/"):
            path = path[:-1]
        return path

    def _path_template_to_regex(self, template: str) -> str:
        """把 `/api/{id}` 形式的模板路径转换为可匹配真实 URL 的正则。"""
        escaped = re.escape(template)
        escaped = re.sub(r"\\\{[^}]+\\\}", r"[^/]+", escaped)
        return f"^{escaped}$"
    
    # ============== 三态资产关联 ==============
    
    async def create_tri_state_asset(
        self,
        runtime_assets: Optional[List[RuntimeAsset]] = None,
        dev_assets: Optional[List[DevAsset]] = None,
        design_assets: Optional[List[DesignAsset]] = None
    ) -> TriStateAsset:
        """创建三态聚合资产。"""
        asset_id = f"tri_{uuid.uuid4().hex[:8]}"
        
        asset = TriStateAsset(
            id=asset_id,
            runtime_assets=runtime_assets or [],
            dev_assets=dev_assets or [],
            design_assets=design_assets or []
        )
        
        await self._repository.save_tri_state_asset(asset)
        
        return asset
    
    async def get_tri_state_asset(self, asset_id: str) -> Optional[TriStateAsset]:
        """读取单个三态聚合资产。"""
        return await self._repository.get_tri_state_asset(asset_id)
    
    async def link_assets(
        self,
        runtime_asset_id: str,
        dev_asset_id: str,
        design_asset_id: Optional[str] = None
    ) -> bool:
        """
        关联三态资产
        
        Args:
            runtime_asset_id: 运行态资产ID
            dev_asset_id: 开发态资产ID
            design_asset_id: 设计态资产ID（可选）
            
        Returns:
            是否成功
        """
        runtime = await self._repository.get_runtime_asset(runtime_asset_id)
        dev = await self._repository.get_dev_asset(dev_asset_id)
        
        if not runtime or not dev:
            return False
        
        # 创建一个新的聚合资产，把运行态和开发态资产先绑定到同一对象上。
        tri_asset = await self.create_tri_state_asset(
            runtime_assets=[runtime],
            dev_assets=[dev]
        )
        
        if design_asset_id:
            design = await self._repository.get_design_asset(design_asset_id)
            if design:
                tri_asset.design_assets.append(design)
        
        # relationships 里保留轻量 ID 关系，便于前端关系图直接消费。
        tri_asset.relationships[runtime_asset_id] = [dev_asset_id]
        if design_asset_id:
            tri_asset.relationships[runtime_asset_id].append(design_asset_id)

        tri_asset.updated_at = datetime.utcnow()
        await self._repository.save_tri_state_asset(tri_asset)
        
        logger.info(
            "assets_linked",
            runtime_id=runtime_asset_id,
            dev_id=dev_asset_id,
            design_id=design_asset_id
        )
        
        return True


# 全局实例
asset_service = AssetService()
